
import os 
import glob 
import tkinter as tk 
from tkinter import messagebox 
import customtkinter as ctk

import subprocess 
import win32api 
import win32con 
import webbrowser 
import configparser 
import time 
import pyperclip 
import pyautogui 
import openpyxl
import re
import threading
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
from urllib.parse import quote
import base64
# .xls 파일 지원을 위한 xlrd (선택적)
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
# Windows COM을 통한 Excel 읽기 (선택적)
try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False 
# PyMuPDF import with error handling
try:
    import fitz # PyMuPDF  # pyright: ignore[reportMissingImports]
    FITZ_AVAILABLE = True
except ImportError:
    FITZ_AVAILABLE = False
    print("경고: PyMuPDF가 설치되지 않았습니다. PDF 인쇄 기능을 사용하려면 'pip install PyMuPDF'를 실행하세요.") 
import sys 
from pywinauto import findwindows 

# Import Selenium components 
from selenium import webdriver 
from selenium.webdriver.common.by import By 
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC 
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException 

# Import for Chrome driver management 
from selenium.webdriver.chrome.service import Service as ChromeService 
from webdriver_manager.chrome import ChromeDriverManager 

# PyInstaller로 패키징된 경우 임시 경로를 사용 
if getattr(sys, 'frozen', False): 
    # _MEIPASS는 PyInstaller가 생성하는 임시 디렉토리 경로 
    application_path = sys._MEIPASS 
else: 
    application_path = os.path.dirname(os.path.abspath(__file__)) 

PDF_TO_PRINTER = os.path.join(application_path, "PDFtoPrinter.exe") 

# PDF 인쇄용 PDFtoPrinter.exe 경로 
PDF_TO_PRINTER = r"C:\Tools\PDFtoPrinter\PDFtoPrinter.exe" 
PRINTER_NAME = None # None이면 기본 프린터 사용 

# config.ini 파일은 사용자의 Documents 폴더에 저장
# C:\Users\USER\Documents\작업환경측정결과 보고서 처리\config.ini
DOCUMENTS_FOLDER = os.path.join(os.path.expanduser("~"), "Documents")
CONFIG_FOLDER = os.path.join(DOCUMENTS_FOLDER, "작업환경측정결과 보고서 처리")
CONFIG_FILE = os.path.join(CONFIG_FOLDER, "config.ini")

# config 폴더가 없으면 생성
if not os.path.exists(CONFIG_FOLDER):
    try:
        os.makedirs(CONFIG_FOLDER)
        print(f"[INFO] 설정 폴더 생성: {CONFIG_FOLDER}")
    except Exception as e:
        print(f"[ERROR] 설정 폴더 생성 실패: {e}")

# 엑셀 파일 경로 동적 설정
# 프로그램 실행 경로 (exe 실행 시 exe 위치, 파이썬 실행 시 스크립트 위치)
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def find_latest_excel_path():
    """ "measurement_business.xlsx" 파일 찾기 """
    # 기본 경로 설정
    target_path = os.path.join(r"Z:\data\측정팀\자동화 툴\MES 프로그램 DB", "measurement_business.xlsx")
    
    if os.path.exists(target_path):
        print(f"[INFO] 엑셀 파일 로드: {target_path}")
        return target_path
    else:
        # 파일이 없을 경우 경고 메시지 출력
        print(f"[WARNING] '{target_path}' 파일을 찾을 수 없습니다.")
        return None

# 초기 실행 시 기존 방식대로 로드
EXCEL_FILE = find_latest_excel_path()

# 네이버 메일 고정 계정 정보
# ⚠️ 중요: 네이버 메일 비밀번호는 네이버 로그인 비밀번호가 아닌 "애플리케이션 비밀번호"를 사용해야 합니다!
# 
# [애플리케이션 비밀번호 생성 방법]
# 1. 네이버ID 접속: https://nid.naver.com
# 2. 네이버ID > 보안설정 > 기본보안설정 > 2단계 인증 > 관리
# 3. 네이버 로그인 비밀번호 재확인
# 4. '애플리케이션 비밀번호 관리' 기능 확인
# 5. 사용하려는 애플리케이션 종류 선택 또는 직접 입력 (예: "SMTP 메일 발송")
# 6. '생성하기' 버튼 클릭
# 7. 생성된 비밀번호를 복사 (한 번 생성 후 다시 확인할 수 없으므로 반드시 복사!)
# 8. GUI의 네이버 PW 입력란에 생성된 비밀번호 입력
#
# ※ 참고: 2025년 6월 24일부터 POP3/IMAP/SMTP 비밀번호 정책이 변경되어 
#          애플리케이션 비밀번호 사용이 필수입니다.
NAVER_EMAIL_ID = "hangyeol5678882@naver.com"
NAVER_EMAIL_PW = "6Y7U3HW2YK4Q"


# 엑셀 파일 경로 확인 함수
def check_excel_file():
    """엑셀 파일 존재 여부 확인"""
    if EXCEL_FILE is None or not os.path.exists(EXCEL_FILE):
        print(f"[오류] 엑셀 파일이 존재하지 않습니다: {EXCEL_FILE}")
        return False
    print(f"[확인] 엑셀 파일 경로: {EXCEL_FILE}")
    return True 

class ReportPrinterApp(ctk.CTk): 
    def __init__(self): 
        super().__init__() 
        self.title("작업환경측정결과 보고서 처리") 
        self.geometry("1280x950")
        # 시작 직후 최대화 (안정적인 동작을 위해 0.1초 후 실행)
        self.after(100, lambda: self.state('zoomed'))
        
        # 최소 창 크기 설정 (너무 작아지지 않도록)
        self.minsize(1000, 700)

        # 아이콘 설정
        try:
            icon_path = os.path.join(application_path, "frog_fixed.ico")
            if os.path.exists(icon_path):
                self.iconbitmap(icon_path)
            else:
                # 개발 환경에서 파일이 직접 옆에 있는 경우 대비
                current_dir_icon = os.path.join(os.path.dirname(os.path.abspath(__file__)), "frog_fixed.ico")
                if os.path.exists(current_dir_icon):
                    self.iconbitmap(current_dir_icon)
        except Exception as e:
            print(f"[DEBUG] 아이콘 설정 실패: {e}")
        
        # 메인 윈도우 그리드 설정 (반응형 레이아웃)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        # 테마 및 색상 설정 (부드럽고 모던한 느낌)
        # 테마 및 색상 설정 (클래식/기본 스타일)
        ctk.set_appearance_mode("System")  # 시스템 기본 설정 (Light/Dark)
        ctk.set_default_color_theme("blue")  # 기본 블루 테마
        
        # 폰트 설정
        self.main_font = ("Malgun Gothic", 12)
        self.header_font = ("Malgun Gothic", 14, "bold")
        
        self.config = configparser.ConfigParser() 
        self.load_config()

        
        # 엑셀 데이터 캐시 (성능 개선)
        self.excel_cache = None
        self.excel_cache_loaded = False
        self.excel_cache_loading = False  # 로딩 중 플래그
        self.excel_app = None  # Windows COM Excel 애플리케이션 재사용
        
        # 프로그램 시작 시 엑셀 파일 미리 로드 (백그라운드)
        self._preload_excel_cache() 

        # ===== 메인 레이아웃 컨테이너 =====
        self.main_frame = ctk.CTkFrame(self)
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # 그리드 설정 (반응형 레이아웃)
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(1, weight=0) # 업체 목록 영역 (고정 크기)
        self.main_frame.grid_rowconfigure(2, weight=1) # 하단 버튼 및 로그 영역 (남은 공간 차지)
        
        # ===== [1] 상단 설정 영역 (년도/반기, 계정 정보) =====
        top_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        top_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        # 1-1. 년도 및 반기 선택 (좌측)
        date_frame = ctk.CTkFrame(top_frame)
        date_frame.pack(side="left", fill="y", padx=(0, 10))
        
        ctk.CTkLabel(date_frame, text="작업 기간 설정", font=self.header_font).pack(pady=5, padx=10, anchor="w")
        
        date_inner = ctk.CTkFrame(date_frame, fg_color="transparent")
        date_inner.pack(padx=10, pady=5)
        
        # 설정에서 저장된 값 불러오기
        default_year = self.config.get("SETTINGS", "Year", fallback="2026")
        default_semester = self.config.get("SETTINGS", "Semester", fallback="상반기")

        ctk.CTkLabel(date_inner, text="년도:", font=self.main_font).pack(side="left", padx=5)
        self.year_entry = ctk.CTkEntry(date_inner, width=60, font=self.main_font)
        self.year_entry.insert(0, default_year)
        self.year_entry.pack(side="left", padx=5)
        
        self.semester = tk.StringVar(value=default_semester)
        ctk.CTkRadioButton(date_inner, text="상반기", variable=self.semester, value="상반기", font=self.main_font).pack(side="left", padx=10)
        ctk.CTkRadioButton(date_inner, text="하반기", variable=self.semester, value="하반기", font=self.main_font).pack(side="left", padx=10)
        
        # 보조 기능: 엑셀 새로고침 버튼 추가 (DB Update)
        self.btn_refresh = ctk.CTkButton(date_inner, text="DB Update", command=self.on_refresh_excel, width=100, height=30, font=("Malgun Gothic", 11, "bold"))
        self.btn_refresh.pack(side="left", padx=(15, 0))
        
        # 1-2. 계정 정보 설정 (K2B) - 좌측 (작업 기간 옆)
        k2b_frame = ctk.CTkFrame(top_frame)
        k2b_frame.pack(side="left", fill="y", padx=(0, 10))
        
        # 타이틀을 요청대로 "계정 정보 설정"으로 유지
        ctk.CTkLabel(k2b_frame, text="계정 정보 설정", font=self.header_font).pack(pady=5, padx=10, anchor="w")
        
        k2b_inner = ctk.CTkFrame(k2b_frame, fg_color="transparent")
        k2b_inner.pack(padx=10, pady=5, fill="x")
        
        ctk.CTkLabel(k2b_inner, text="K2B ID:", font=self.main_font).pack(side="left", padx=5)
        self.k2b_ID_entry = ctk.CTkEntry(k2b_inner, width=120, font=self.main_font)
        self.k2b_ID_entry.pack(side="left", padx=5)
        self.k2b_ID_entry.insert(0, self.config.get("K2B", "ID", fallback=""))
        
        ctk.CTkLabel(k2b_inner, text="PW:", font=self.main_font).pack(side="left", padx=5)
        self.k2b_PW_entry = ctk.CTkEntry(k2b_inner, show="*", width=120, font=self.main_font)
        self.k2b_PW_entry.pack(side="left", padx=5)
        self.k2b_PW_entry.insert(0, self.config.get("K2B", "PW", fallback=""))
        
        btn_k2b_connect = ctk.CTkButton(k2b_inner, text="K2B 전송", command=self.connect_to_k2b, width=100, height=30, font=("Malgun Gothic", 11, "bold"))
        btn_k2b_connect.pack(side="left", padx=10)
        
        # 1-3. 보고서 발송 전용 계정 (Naver) - 우측 끝으로 이동
        naver_frame = ctk.CTkFrame(top_frame)
        naver_frame.pack(side="right", fill="y")
        
        ctk.CTkLabel(naver_frame, text="보고서 발송 전용 계정", font=self.header_font).pack(pady=5, padx=10, anchor="w")
        
        naver_inner = ctk.CTkFrame(naver_frame, fg_color="transparent")
        naver_inner.pack(padx=10, pady=5, fill="x")
        
        ctk.CTkLabel(naver_inner, text="네이버 ID:", font=self.main_font).pack(side="left", padx=5)
        self.naver_ID_entry = ctk.CTkEntry(naver_inner, width=120, font=self.main_font)
        self.naver_ID_entry.pack(side="left", padx=5)
        self.naver_ID_entry.insert(0, NAVER_EMAIL_ID)
        self.naver_ID_entry.configure(state="readonly")
        
        ctk.CTkLabel(naver_inner, text="앱 비밀번호:", font=self.main_font).pack(side="left", padx=5)
        self.naver_PW_entry = ctk.CTkEntry(naver_inner, show="*", width=150, font=self.main_font)
        self.naver_PW_entry.pack(side="left", padx=5)
        self.naver_PW_entry.insert(0, self.config.get("NAVER_EMAIL", "PW", fallback=""))

        # ===== [2] 업체명 입력 영역 =====
        self.entries = [] 
        self.k2b_status_labels = []
        self.path_labels = []
        self.email_labels = []
        self.email_status_labels = []
        self.listbox_widgets = {}
        
        self.company_frame = ctk.CTkFrame(self.main_frame)
        self.company_frame.grid(row=1, column=0, sticky="nsew", pady=10)
        
        ctk.CTkLabel(self.company_frame, text="업체 목록 및 처리 현황", font=self.header_font).pack(pady=10, padx=10, anchor="w")
        
        # 스크롤 가능한 프레임 사용 (혹은 그냥 프레임) - 10개 고정이므로 일반 프레임 사용
        self.company_inner_frame = ctk.CTkFrame(self.company_frame, fg_color="transparent")
        self.company_inner_frame.pack(fill="x", expand=False, padx=10, pady=5)

        # 헤더
        header_frame = ctk.CTkFrame(self.company_inner_frame, height=30, fg_color="transparent")
        header_frame.pack(fill="x", pady=(0, 5))
        ctk.CTkLabel(header_frame, text="업체명", width=180, anchor="w", font=("Malgun Gothic", 11, "bold")).pack(side="left", padx=(45, 5))
        ctk.CTkLabel(header_frame, text="K2B 상태", width=60, anchor="center", font=("Malgun Gothic", 11, "bold")).pack(side="left", padx=5)
        ctk.CTkLabel(header_frame, text="폴더 경로", width=500, anchor="w", font=("Malgun Gothic", 11, "bold")).pack(side="left", padx=10)
        ctk.CTkLabel(header_frame, text="이메일", width=200, anchor="w", font=("Malgun Gothic", 11, "bold")).pack(side="left", padx=10)
        ctk.CTkLabel(header_frame, text="발송 상태", width=60, anchor="center", font=("Malgun Gothic", 11, "bold")).pack(side="left", padx=5)

        for i in range(10): 
            self.add_entry_row(self.company_inner_frame, i)

        # ===== [3] 하단 버튼 및 로그 영역 =====
        bottom_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        bottom_frame.grid(row=2, column=0, sticky="nsew", pady=(10, 0))
        
        # 버튼 영역
        button_frame = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=(0, 10))
        
        btn_print = ctk.CTkButton(button_frame, text="보고서 출력(P)", width=140, height=40, command=self.on_print, font=self.header_font)
        btn_print.pack(side="left", padx=10)
        self.bind("<Alt-p>", lambda e: self.on_print()) 

        btn_send_email = ctk.CTkButton(button_frame, text="메일 발송(M)", width=140, height=40, command=self.send_emails_manual, font=self.header_font, fg_color="#2E7D32", hover_color="#1B5E20")
        btn_send_email.pack(side="left", padx=10)
        self.bind("<Alt-m>", lambda e: self.send_emails_manual()) 

        btn_reset = ctk.CTkButton(button_frame, text="초기화(I)", width=120, height=40, command=self.reset_all, font=self.header_font, fg_color="#757575", hover_color="#616161")
        btn_reset.pack(side="left", padx=10)
        self.bind("<Alt-i>", lambda e: self.reset_all()) 

        btn_close = ctk.CTkButton(button_frame, text="닫기(C)", width=100, height=40, command=self.on_close, font=self.header_font, fg_color="#C62828", hover_color="#B71C1C")
        btn_close.pack(side="right", padx=10)
        self.bind("<Alt-c>", lambda e: self.on_close()) 

        self.protocol("WM_DELETE_WINDOW", self.on_close) 

        # 로그 영역
        log_frame = ctk.CTkFrame(bottom_frame)
        log_frame.pack(fill="both", expand=True)
        
        ctk.CTkLabel(log_frame, text="작업 로그", font=("Malgun Gothic", 11, "bold")).pack(anchor="w", padx=10, pady=(5,0))

        self.result_text = ctk.CTkTextbox(log_frame, height=150, font=("Consolas", 10))
        self.result_text.pack(fill="both", expand=True, padx=10, pady=5) 
 

 

    def load_config(self): 
        """설정 파일을 로드합니다.""" 
        if os.path.exists(CONFIG_FILE): 
            self.config.read(CONFIG_FILE) 
            
    def save_config(self): 
        """설정 파일을 저장합니다.""" 
        if not self.config.has_section("K2B"): 
            self.config.add_section("K2B") 
        self.config.set("K2B", "ID", self.k2b_ID_entry.get()) 
        self.config.set("K2B", "PW", self.k2b_PW_entry.get()) 
        
        # 년도 및 반기 설정 저장
        if not self.config.has_section("SETTINGS"):
            self.config.add_section("SETTINGS")
        self.config.set("SETTINGS", "Year", self.year_entry.get())
        self.config.set("SETTINGS", "Semester", self.semester.get())

        # 네이버 메일 비밀번호 저장 (ID는 고정값)
        if not self.config.has_section("NAVER_EMAIL"):
            self.config.add_section("NAVER_EMAIL")
        self.config.set("NAVER_EMAIL", "PW", self.naver_PW_entry.get())
        
        with open(CONFIG_FILE, "w") as f: 
            self.config.write(f) 

    def on_close(self): 
        """프로그램 종료 시 호출되어 설정을 저장하고 창을 닫습니다.""" 
        self.save_config()
        # Excel 애플리케이션 종료
        if self.excel_app:
            try:
                self.excel_app.Quit()
            except:
                pass
        self.destroy()

    def on_refresh_excel(self):
        """엑셀 파일을 새로고침하여 최신 데이터를 반영합니다."""
        global EXCEL_FILE
        
        if self.excel_cache_loading:
            self.log_result("[경고] 현재 데이터를 로드 중입니다. 잠시 후 다시 시도해 주세요.")
            return

        self.log_result("[INFO] DB Update (데이터 동기화)를 시작합니다...")
        
        # 1. 최신 엑셀 파일 경로 다시 찾기
        new_path = find_latest_excel_path()
        if new_path:
            EXCEL_FILE = new_path
            
            # 2. 캐시 초기화
            self.excel_cache_loaded = False
            
            # 3. 백그라운드에서 데이터 다시 로드
            self._preload_excel_cache()
            self.log_result(f"[확인] 최신 엑셀 파일 경로를 갱신했습니다: {os.path.basename(EXCEL_FILE)}")
        else:
            self.log_result("[오류] 최신 엑셀 파일을 찾을 수 없습니다. 경로를 확인해 주세요.")
    
    def _preload_excel_cache(self):
        """프로그램 시작 시 백그라운드에서 엑셀 파일을 미리 로드"""
        def load_in_background():
            self._load_excel_cache()
        
        thread = threading.Thread(target=load_in_background, daemon=True)
        thread.start()
        print("[DEBUG] 엑셀 파일 백그라운드 로딩 시작...")
    
    def _load_excel_cache(self):
        """엑셀 파일을 한 번만 로드하여 메모리에 캐싱"""
        if self.excel_cache_loaded:
            return self.excel_cache
        
        # 이미 로딩 중이면 대기
        if self.excel_cache_loading:
            max_wait = 10  # 최대 10초 대기
            wait_time = 0
            while self.excel_cache_loading and wait_time < max_wait:
                time.sleep(0.1)
                wait_time += 0.1
            if self.excel_cache_loaded:
                return self.excel_cache
            return None
        
        self.excel_cache_loading = True
        
        try:
            if EXCEL_FILE is None or not os.path.exists(EXCEL_FILE):
                return None
            
            file_ext = os.path.splitext(EXCEL_FILE)[1].lower()
            
            # .xlsx 파일 처리 (openpyxl 사용)
            if file_ext == '.xlsx':
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
                    ws = wb.active
                    cache_data = []
                    
                    # iter_rows를 통해 데이터 읽기 (min_row=2는 헤더 제외)
                    # values_only=True를 사용하면 튜플 형태로 값만 반환됨
                    # 엑셀 열 인덱스 (0-based in tuple): B=1, C=2, D=3 ... BL=63
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row:
                            continue
                            
                        try:
                            # 튜플 범위 확인
                            if len(row) <= 3:
                                continue
                                
                            year_val = str(row[1]).strip() if row[1] is not None else ""
                            semester_val = str(row[2]).strip() if row[2] is not None else ""
                            company_val = str(row[3]).strip() if row[3] is not None else ""
                            
                            email_val = ""
                            if len(row) > 63:
                                email_cell = row[63]
                                email_val = str(email_cell).strip() if email_cell is not None else ""
                                # 이메일 주소만 추출
                                email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                                email_matches = re.findall(email_pattern, email_val)
                                if email_matches:
                                    email_val = email_matches[0]
                                    
                            cache_data.append({
                                'year': year_val,
                                'semester': semester_val,
                                'company': company_val,
                                'email': email_val
                            })
                        except Exception as inner_e:
                            continue
                            
                    wb.close()
                    self.excel_cache = cache_data
                    self.excel_cache_loaded = True
                    self.excel_cache_loading = False
                    print(f"[DEBUG] 엑셀 데이터 캐시 로드 완료 (openpyxl): {len(cache_data)}개 행")
                    return cache_data
                    
                except Exception as e:
                    print(f"[DEBUG] openpyxl 로드 실패: {e}")
                    # 실패 시 아래 다른 방법 시도하지 않고 종료 (xlsx는 openpyxl이 담당)
                    self.excel_cache_loading = False
                    return None
            
            if file_ext == '.xls' and WIN32COM_AVAILABLE:
                # Windows COM을 사용하여 데이터 로드 (다중 사용자 환경 대응)
                max_retries = 3
                retry_delay = 0.5  # 초
                
                for attempt in range(max_retries):
                    try:
                        if not self.excel_app:
                            self.excel_app = win32com.client.Dispatch("Excel.Application")
                            # 엑셀이 화면에 나타나지 않도록 즉시 설정
                            self.excel_app.Visible = False
                            self.excel_app.DisplayAlerts = False
                            self.excel_app.ScreenUpdating = False  # 화면 갱신 차단
                        
                        # 읽기 전용으로 파일 열기 (다중 사용자 환경 대응)
                        # ReadOnly=True: 읽기 전용으로 열기
                        # UpdateLinks=0: 외부 링크 업데이트 안 함
                        # IgnoreReadOnlyRecommended=True: 읽기 전용 권장 무시
                        workbook = self.excel_app.Workbooks.Open(
                            EXCEL_FILE, 
                            ReadOnly=True, 
                            UpdateLinks=0,
                            IgnoreReadOnlyRecommended=True
                        )
                        worksheet = workbook.ActiveSheet
                        last_row = worksheet.UsedRange.Rows.Count
                        
                        # 모든 데이터를 리스트로 로드
                        cache_data = []
                        # 전체 데이터를 한 번에 튜플로 가져옴 (성능 최적화 핵심)
                        # A1부터 BL(64)열까지의 모든 범위를 한 번에 메모리로 로드
                        all_values = worksheet.Range(worksheet.Cells(1, 1), worksheet.Cells(last_row, 64)).Value
                        
                        # all_values는 (행, 열) 형태의 중첩 튜플 (0-indexed)
                        # 1행은 헤더이므로 인덱스 1부터 시작
                        if all_values and len(all_values) > 1:
                            for row_idx in range(1, len(all_values)):
                                row_data = all_values[row_idx]
                                try:
                                    # row_data[1]: B열, [2]: C열, [3]: D열, [63]: BL열
                                    year_val = str(row_data[1]).strip() if row_data[1] is not None else ""
                                    semester_val = str(row_data[2]).strip() if row_data[2] is not None else ""
                                    company_val = str(row_data[3]).strip() if row_data[3] is not None else ""
                                    
                                    email_val = ""
                                    if len(row_data) >= 64:
                                        email_cell = row_data[63]
                                        email_val = str(email_cell).strip() if email_cell is not None else ""
                                        # 이메일 주소만 추출
                                        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                                        email_matches = re.findall(email_pattern, email_val)
                                        if email_matches:
                                            email_val = email_matches[0]
                                    
                                    cache_data.append({
                                        'year': year_val,
                                        'semester': semester_val,
                                        'company': company_val,
                                        'email': email_val
                                    })
                                except:
                                    continue
                        
                        workbook.Close(False)
                        self.excel_cache = cache_data
                        self.excel_cache_loaded = True
                        self.excel_cache_loading = False
                        print(f"[DEBUG] 엑셀 데이터 캐시 로드 완료: {len(cache_data)}개 행")
                        return cache_data
                    except Exception as e:
                        error_msg = str(e).lower()
                        # 파일이 사용 중이거나 잠겨있는 경우 재시도
                        if any(keyword in error_msg for keyword in ["파일이 사용 중", "file is in use", "locked", "permission", "액세스"]):
                            if attempt < max_retries - 1:
                                print(f"[DEBUG] 파일이 사용 중입니다. {retry_delay}초 후 재시도... ({attempt + 1}/{max_retries})")
                                time.sleep(retry_delay)
                                retry_delay *= 2  # 지수 백오프
                                continue
                            else:
                                print(f"[DEBUG] 파일이 계속 사용 중입니다. xlrd로 시도합니다.")
                                break
                        else:
                            print(f"[DEBUG] Windows COM으로 캐시 로드 실패: {e}")
                            break
            
            # Windows COM이 실패하면 xlrd 시도 (다중 사용자 환경에 더 안전함)
            if file_ext == '.xls' and XLRD_AVAILABLE:
                max_retries = 3
                retry_delay = 0.5
                
                for attempt in range(max_retries):
                    try:
                        # xlrd는 기본적으로 읽기 전용이며, 파일 잠금에 더 안전함
                        # on_demand=True: 필요한 시트만 메모리에 로드
                        book = xlrd.open_workbook(EXCEL_FILE, on_demand=True)
                        sheet = book.sheet_by_index(0)
                        cache_data = []
                        
                        for row_idx in range(1, sheet.nrows):
                            try:
                                year_val = str(sheet.cell_value(row_idx, 1)).strip() if sheet.cell_value(row_idx, 1) else ""
                                semester_val = str(sheet.cell_value(row_idx, 2)).strip() if sheet.cell_value(row_idx, 2) else ""
                                company_val = str(sheet.cell_value(row_idx, 3)).strip() if sheet.cell_value(row_idx, 3) else ""
                                
                                email_val = ""
                                if sheet.ncols > 63:
                                    email_val = str(sheet.cell_value(row_idx, 63)).strip() if sheet.cell_value(row_idx, 63) else ""
                                    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                                    email_matches = re.findall(email_pattern, email_val)
                                    if email_matches:
                                        email_val = email_matches[0]
                                
                                cache_data.append({
                                    'year': year_val,
                                    'semester': semester_val,
                                    'company': company_val,
                                    'email': email_val
                                })
                            except:
                                continue
                        
                        self.excel_cache = cache_data
                        self.excel_cache_loaded = True
                        self.excel_cache_loading = False
                        print(f"[DEBUG] 엑셀 데이터 캐시 로드 완료 (xlrd): {len(cache_data)}개 행")
                        return cache_data
                    except Exception as e:
                        error_msg = str(e).lower()
                        # 파일이 사용 중이거나 잠겨있는 경우 재시도
                        if any(keyword in error_msg for keyword in ["file is in use", "locked", "permission", "액세스", "사용 중"]):
                            if attempt < max_retries - 1:
                                print(f"[DEBUG] 파일이 사용 중입니다. {retry_delay}초 후 재시도... ({attempt + 1}/{max_retries})")
                                time.sleep(retry_delay)
                                retry_delay *= 2
                                continue
                            else:
                                print(f"[DEBUG] 파일이 계속 사용 중입니다. 캐시 로드 실패.")
                                return None
                        else:
                            print(f"[DEBUG] xlrd로 캐시 로드 실패: {e}")
                            return None
            
            self.excel_cache_loading = False
            return None
        except Exception as e:
            print(f"[DEBUG] 엑셀 캐시 로드 오류: {e}")
            self.excel_cache_loading = False
            return None
    
    def find_email_from_excel(self, year, semester, company_name):
        """
        엑셀 파일에서 년도, 구분(반기), 사업장명과 일치하는 행의 메일 주소(BL열)를 찾습니다.
        캐시된 데이터를 사용하여 빠르게 검색하며, 정확한 매칭을 우선순위로 둡니다.
        
        Args:
            year: 년도 (문자열)
            semester: 반기 ("상반기" 또는 "하반기")
            company_name: 사업장명 (문자열)
        
        Returns:
            메일 주소 (문자열), 없으면 None
        """
        try:
            # 캐시된 데이터 로드 (한 번만)
            cache_data = self._load_excel_cache()
            if not cache_data:
                return None
            
            # 정규화 함수
            def normalize_company_name(name):
                """사업장명을 정규화 (공백, 괄호 제거)"""
                normalized = re.sub(r'[()（）\s]', '', name)
                return normalized.lower()
            
            norm_target = normalize_company_name(company_name)
            
            best_email = None
            best_score = 0
            best_company_val = ""
            
            # 캐시된 데이터에서 검색
            for row in cache_data:
                year_val = row['year']
                semester_val = row['semester']
                company_val = row['company']
                email_val = row['email']
                
                # 년도 매칭
                year_match = False
                if year_val == str(year):
                    year_match = True
                elif year_val.endswith(str(year)) or str(year) in year_val:
                    year_match = True
                
                # "번외"가 포함된 데이터는 배제
                if "번외" in company_val:
                    continue
                
                # 구분 매칭
                semester_match = False
                if semester_val == semester:
                    semester_match = True
                elif semester == "상반기" and ("상" in semester_val or "1" in semester_val):
                    semester_match = True
                elif semester == "하반기" and ("하" in semester_val or "2" in semester_val):
                    semester_match = True
                
                if not (year_match and semester_match):
                    continue
                
                # 사업장명 매칭 및 점수 산정
                norm_val = normalize_company_name(company_val)
                score = 0
                
                # 1. 정확히 일치 (가장 높은 우선순위)
                if norm_target == norm_val:
                    score = 3
                # 2. 엑셀 값이 검색어(폴더명)를 포함 (검색어 '중앙', 엑셀 '중앙대학교')
                elif norm_target in norm_val:
                    score = 2
                # 3. 검색어(폴더명)가 엑셀 값을 포함 (검색어 '중앙대학교(본관)', 엑셀 '중앙대학교')
                #    유사한 다른 업체(짧은 이름)가 매칭될 위험이 있어 점수가 낮음
                elif norm_val in norm_target:
                    score = 1
                
                if score > best_score:
                    best_score = score
                    best_email = email_val if email_val else None
                    best_company_val = company_val
                    # 정확한 매칭을 찾으면 즉시 반환 (최적화)
                    if score == 3:
                        break
            
            if best_email:
                print(f"[DEBUG] 이메일 찾기 성공: 검색어='{company_name}' -> 매칭='{best_company_val}' (점수: {best_score}) -> 이메일='{best_email}'")
            elif best_score > 0 and not best_email:
                print(f"[DEBUG] 매칭된 업체는 찾았으나 이메일이 없음: {best_company_val}")
                
            return best_email
            
        except Exception as e:
            print(f"[DEBUG] 엑셀 검색 오류: {e}")
            return None
    
    def _find_email_from_xls(self, year, semester, company_name):
        """
        xlrd를 사용하여 .xls 파일에서 메일 주소를 찾습니다.
        """
        try:
            book = xlrd.open_workbook(EXCEL_FILE)
            sheet = book.sheet_by_index(0)
            
            print(f"[DEBUG] xlrd로 엑셀 파일 열기 성공. 행 수: {sheet.nrows}, 열 수: {sheet.ncols}")
            print(f"[DEBUG] 검색 조건: 년도={year}, 구분={semester}, 사업장명={company_name}")
            
            # BL열 인덱스 계산: B=2, L=12이므로 BL는 2*26 + 12 = 64번째 열 (인덱스 63)
            bl_col_index = 63  # xlrd 기준 인덱스 63 (BL열)
            
            if sheet.ncols < 64:
                print(f"[DEBUG] 경고: 최대 열 수가 {sheet.ncols}입니다.")
                bl_col_index = min(63, sheet.ncols - 1)
            
            # B열(인덱스 1): 년도, C열(인덱스 2): 구분, D열(인덱스 3): 사업장명, BL열: 메일
            for row_idx in range(1, sheet.nrows):  # 0행은 헤더로 가정
                try:
                    year_val = str(sheet.cell_value(row_idx, 1)).strip() if sheet.cell_value(row_idx, 1) else ""  # B열
                    semester_val = str(sheet.cell_value(row_idx, 2)).strip() if sheet.cell_value(row_idx, 2) else ""  # C열
                    company_val = str(sheet.cell_value(row_idx, 3)).strip() if sheet.cell_value(row_idx, 3) else ""  # D열
                    
                    # BL열 값 가져오기
                    email_val = ""
                    if sheet.ncols > bl_col_index:
                        email_val = str(sheet.cell_value(row_idx, bl_col_index)).strip() if sheet.cell_value(row_idx, bl_col_index) else ""
                        # 이메일 주소만 추출 (이메일 형식이 포함된 부분만)
                        import re
                        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                        email_matches = re.findall(email_pattern, email_val)
                        if email_matches:
                            email_val = email_matches[0]  # 첫 번째 이메일 주소만 사용
                    
                    # 디버깅: 처음 몇 개 행만 출력
                    if row_idx <= 5:
                        print(f"[DEBUG] Row {row_idx+1}: B={year_val}, C={semester_val}, D={company_val[:20]}..., BL={email_val[:30] if email_val else 'None'}...")
                    
                    # 년도 매칭
                    year_match = False
                    if year_val == str(year):
                        year_match = True
                    elif year_val.endswith(str(year)) or str(year) in year_val:
                        year_match = True
                    
                    # 구분 매칭
                    semester_match = False
                    if semester_val == semester:
                        semester_match = True
                    elif semester == "상반기" and ("상" in semester_val or "1" in semester_val):
                        semester_match = True
                    elif semester == "하반기" and ("하" in semester_val or "2" in semester_val):
                        semester_match = True
                    
                    # 사업장명 매칭 (공백, 괄호 등을 제거하고 비교)
                    def normalize_company_name(name):
                        """사업장명을 정규화 (공백, 괄호 제거)"""
                        # 공백 제거, 괄호 제거
                        normalized = re.sub(r'[()（）\s]', '', name)
                        return normalized.lower()
                    
                    company_match = (
                        company_name in company_val or 
                        company_val in company_name or
                        normalize_company_name(company_name) in normalize_company_name(company_val) or
                        normalize_company_name(company_val) in normalize_company_name(company_name)
                    )
                    
                    # 모든 조건이 일치하는지 확인
                    if year_match and semester_match and company_match:
                        print(f"[DEBUG] 매칭 발견! Row {row_idx+1}: {company_val} -> {email_val}")
                        return email_val if email_val else None
                except Exception as e:
                    print(f"[DEBUG] Row {row_idx+1} 처리 중 오류: {e}")
                    continue
            
            print(f"[DEBUG] 매칭된 행이 없습니다. (검색한 행 수: {sheet.nrows - 1})")
            return None
            
        except Exception as e:
            print(f"[DEBUG] xlrd로 엑셀 파일 읽기 오류: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _find_email_from_xls_com(self, year, semester, company_name):
        """
        Windows COM을 사용하여 Excel 애플리케이션을 통해 .xls 파일에서 메일 주소를 찾습니다.
        """
        excel_app = None
        try:
            print(f"[DEBUG] Windows COM을 통해 엑셀 파일 열기 시도...")
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            
            workbook = excel_app.Workbooks.Open(EXCEL_FILE)
            worksheet = workbook.ActiveSheet
            
            print(f"[DEBUG] 엑셀 파일 열기 성공. 행 수: {worksheet.UsedRange.Rows.Count}")
            print(f"[DEBUG] 검색 조건: 년도={year}, 구분={semester}, 사업장명={company_name}")
            
            # BL열 인덱스: B=2, L=12이므로 BL는 64번째 열
            bl_col = 64
            
            # 사용된 범위의 마지막 행 찾기
            last_row = worksheet.UsedRange.Rows.Count
            
            # B열(인덱스 2): 년도, C열(인덱스 3): 구분, D열(인덱스 4): 사업장명, BL열(인덱스 64): 메일
            for row_idx in range(2, last_row + 1):  # 1행은 헤더로 가정
                try:
                    year_val = str(worksheet.Cells(row_idx, 2).Value).strip() if worksheet.Cells(row_idx, 2).Value else ""  # B열
                    semester_val = str(worksheet.Cells(row_idx, 3).Value).strip() if worksheet.Cells(row_idx, 3).Value else ""  # C열
                    company_val = str(worksheet.Cells(row_idx, 4).Value).strip() if worksheet.Cells(row_idx, 4).Value else ""  # D열
                    
                    # BL열 값 가져오기
                    email_val = ""
                    try:
                        email_cell = worksheet.Cells(row_idx, bl_col).Value
                        email_val = str(email_cell).strip() if email_cell else ""
                        # 이메일 주소만 추출 (이메일 형식이 포함된 부분만)
                        import re
                        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                        email_matches = re.findall(email_pattern, email_val)
                        if email_matches:
                            email_val = email_matches[0]  # 첫 번째 이메일 주소만 사용
                    except:
                        pass
                    
                    # 디버깅: 처음 몇 개 행만 출력
                    if row_idx <= 5:
                        print(f"[DEBUG] Row {row_idx}: B={year_val}, C={semester_val}, D={company_val[:20]}..., BL={email_val[:30] if email_val else 'None'}...")
                    
                    # 년도 매칭
                    year_match = False
                    if year_val == str(year):
                        year_match = True
                    elif year_val.endswith(str(year)) or str(year) in year_val:
                        year_match = True
                    
                    # 구분 매칭
                    semester_match = False
                    if semester_val == semester:
                        semester_match = True
                    elif semester == "상반기" and ("상" in semester_val or "1" in semester_val):
                        semester_match = True
                    elif semester == "하반기" and ("하" in semester_val or "2" in semester_val):
                        semester_match = True
                    
                    # 사업장명 매칭 (공백, 괄호 등을 제거하고 비교)
                    def normalize_company_name(name):
                        """사업장명을 정규화 (공백, 괄호 제거)"""
                        # 공백 제거, 괄호 제거
                        normalized = re.sub(r'[()（）\s]', '', name)
                        return normalized.lower()
                    
                    company_match = (
                        company_name in company_val or 
                        company_val in company_name or
                        normalize_company_name(company_name) in normalize_company_name(company_val) or
                        normalize_company_name(company_val) in normalize_company_name(company_name)
                    )
                    
                    # 모든 조건이 일치하는지 확인
                    if year_match and semester_match and company_match:
                        print(f"[DEBUG] 매칭 발견! Row {row_idx}: {company_val} -> {email_val}")
                        workbook.Close(False)
                        excel_app.Quit()
                        return email_val if email_val else None
                except Exception as e:
                    print(f"[DEBUG] Row {row_idx} 처리 중 오류: {e}")
                    continue
            
            print(f"[DEBUG] 매칭된 행이 없습니다. (검색한 행 수: {last_row - 1})")
            workbook.Close(False)
            excel_app.Quit()
            return None
            
        except Exception as e:
            print(f"[DEBUG] Windows COM으로 엑셀 파일 읽기 오류: {e}")
            import traceback
            traceback.print_exc()
            if excel_app:
                try:
                    excel_app.Quit()
                except:
                    pass
            return None 
        
    def add_entry_row(self, parent, i): 
        row_frame = ctk.CTkFrame(parent, fg_color="transparent")
        row_frame.pack(fill="x", pady=2)
        
        ctk.CTkLabel(row_frame, text=f"{i+1}", width=30, anchor="center", font=self.main_font).pack(side="left", padx=5)
        
        entry = ctk.CTkEntry(row_frame, width=180, font=self.main_font)
        entry.pack(side="left", padx=5)
        self.entries.append(entry) 
        entry.bind("<Return>", lambda e, idx=i: self.focus_next_entry(idx)) 
        entry.bind("<Tab>", lambda e, idx=i: self.focus_next_entry(idx)) 
        
        # K2B 전송 결과 상태 표시 레이블
        k2b_status_label = ctk.CTkLabel(row_frame, text="대기", width=60, anchor="center", fg_color="#E0E0E0", text_color="black", corner_radius=5, font=("Malgun Gothic", 10))
        k2b_status_label.pack(side="left", padx=5)
        self.k2b_status_labels.append(k2b_status_label)
        
        # 경로 표시 (긴 경로로 인한 레이아웃 깨짐 방지를 위해 컨테이너 사용)
        path_container = ctk.CTkFrame(row_frame, width=500, height=25, fg_color="transparent")
        path_container.pack(side="left", padx=10)
        path_container.pack_propagate(False)
        
        path_label = ctk.CTkLabel(path_container, text="", anchor="w", text_color="blue", cursor="hand2", font=("Malgun Gothic", 10))
        path_label.pack(side="left", fill="both", expand=True)
        self.path_labels.append(path_label) 
        path_label.bind("<Button-1>", lambda e, idx=i: self.open_folder(idx))
        
        # 이메일 표시 (긴 이메일로 인한 레이아웃 깨짐 방지를 위해 컨테이너 사용)
        email_container = ctk.CTkFrame(row_frame, width=200, height=30, fg_color="transparent")
        email_container.pack(side="left", padx=10)
        email_container.pack_propagate(False)

        email_label = ctk.CTkLabel(email_container, text="", anchor="w", text_color="green", font=self.main_font)
        email_label.pack(side="left", fill="both", expand=True)
        self.email_labels.append(email_label)
        
        # 메일 발송 상태 표시 레이블
        status_label = ctk.CTkLabel(row_frame, text="대기", width=60, anchor="center", fg_color="#E0E0E0", text_color="black", corner_radius=5, font=("Malgun Gothic", 10))
        status_label.pack(side="left", padx=5)
        self.email_status_labels.append(status_label) 
            
    def focus_next_entry(self, idx): 
        self.destroy_listbox() 
        company_name = self.entries[idx].get().strip() 
        self.path_labels[idx].configure(text="") 
        self.email_labels[idx].configure(text="")  # 메일 주소 초기화

        if not company_name: 
            if idx + 1 < len(self.entries): 
                self.entries[idx + 1].focus_set() 
            return "break" 
        
        year = self.year_entry.get().strip() 
        semester = self.semester.get().strip() 
        
        # 파일 경로 수정: https -> Z: 
        base_path = fr"Z:\data\측정팀\측정보고서\{year}년\{semester}" 
        
        try: 
            # "번외"가 포함된 폴더는 검색에서 제외
            matches = [f for f in os.listdir(base_path) if company_name in f and "번외" not in f] 
        except Exception: 
            messagebox.showerror("오류", f"경로 접근 실패: {base_path}") 
            return "break" 
        
        if len(matches) == 0: 
            messagebox.showwarning("검색 실패", f"'{company_name}' → 해당 폴더 없음") 
            self.entries[idx].focus_set() 
            # 메일 주소 초기화
            self.email_labels[idx].configure(text="")
        elif len(matches) == 1: 
            # 폴더가 하나만 매칭되는 경우에만 메일 주소 검색
            full_path = os.path.join(base_path, matches[0]) 
            self.entries[idx].delete(0, "end") 
            self.entries[idx].insert(0, matches[0]) 
            self.path_labels[idx].configure(text=full_path)
            # 선택된 폴더명으로 메일 주소 검색 (정확한 매칭)
            email = self.find_email_from_excel(year, semester, matches[0])
            if email:
                self.email_labels[idx].configure(text=email)
                print(f"[DEBUG] 메일 주소 표시 (폴더명): {email}")
            else:
                self.email_labels[idx].configure(text="")

                print(f"[DEBUG] 메일 주소를 찾을 수 없습니다: {matches[0]}")
            if idx + 1 < len(self.entries): 
                self.entries[idx + 1].focus_set() 
        else: 
            # 여러 개의 매칭이 있는 경우 리스트박스 표시
            # 리스트박스에서 선택하기 전까지는 메일 주소를 표시하지 않음
            self.email_labels[idx].configure(text="")
            self.show_listbox(matches, idx) 
 
        return "break" 

    def show_listbox(self, options, idx): 
        self.destroy_listbox() 
        
        # entry 위젯 찾기
        entry_widget = self.entries[idx]
        
        # 위치 계산 (메인 윈도우 기준 상대 좌표)
        # update_idletasks()를 호출하여 최신 좌표를 확보
        self.update_idletasks()
        
        x = entry_widget.winfo_rootx() - self.winfo_rootx()
        y = entry_widget.winfo_rooty() - self.winfo_rooty() + entry_widget.winfo_height()
        
        # 리스트박스는 tkinter 위젯 사용 (CustomTkinter에는 기본 리스트박스가 없음)
        # self(메인 윈도우)를 부모로 하여 다른 위젯 위에 뜨도록 함
        
        # 폰트 크기 조정 (작게)
        listbox_font = ("Malgun Gothic", 10)
        
        # 가장 긴 텍스트에 맞춰 너비 조정 (한글 고려하여 넉넉하게 계산)
        max_len = 0
        for opt in options:
            # 한글은 길이를 2로 계산하는 것이 안전
            current_len = 0
            for char in opt:
                if ord(char) > 127: # 한글 등 멀티바이트
                    current_len += 2
                else:
                    current_len += 1
            max_len = max(max_len, current_len)
            
        # 최소 30, 최대 80으로 제한
        width = max(30, min(max_len + 2, 80))
        
        listbox = tk.Listbox(self, selectmode="single", height=min(len(options), 10), width=width, font=listbox_font) 
        
        # place로 위치 지정
        listbox.place(x=x, y=y)
        
        self.listbox_widgets[idx] = listbox 
 
        for opt in options: 
            listbox.insert("end", opt) 
        def on_select(event): 
            try: 
                selected_idx = listbox.curselection()[0] 
                choice = options[selected_idx] 
                self.entries[idx].delete(0, "end") 
                self.entries[idx].insert(0, choice) 
                year = self.year_entry.get().strip() 
                semester = self.semester.get().strip() 
                base_path = fr"Z:\data\측정팀\측정보고서\{year}년\{semester}" 
                full_path = os.path.join(base_path, choice) 
                self.path_labels[idx].configure(text=full_path)

                # 선택된 폴더명으로 메일 주소 검색
                email = self.find_email_from_excel(year, semester, choice)
                if email:
                    self.email_labels[idx].configure(text=email)
                    print(f"[DEBUG] 메일 주소 표시 (리스트박스 선택): {email}")
                else:
                    self.email_labels[idx].configure(text="")
                    print(f"[DEBUG] 메일 주소를 찾을 수 없습니다 (리스트박스): {choice}")

                self.destroy_listbox() 
                if idx + 1 < len(self.entries): 
                    self.entries[idx + 1].focus_set() 
            except IndexError: 
                pass 
        listbox.bind("<Double-1>", on_select) 
        listbox.focus_set() 

    def destroy_listbox(self): 
        widgets_to_destroy = list(self.listbox_widgets.values()) 
        self.listbox_widgets.clear() 
        for widget in widgets_to_destroy: 
            widget.destroy() 

    def open_folder(self, idx): 
        path = self.path_labels[idx].cget("text") 
        if path and os.path.exists(path): 
            try: 
                os.startfile(path) 
                self.log_result(f"[폴더 열기] {path}") 
            except Exception as e: 
                messagebox.showerror("오류", f"폴더를 여는 데 실패했습니다: {path}\n오류: {e}") 
        else: 
            messagebox.showwarning("경로 없음", "유효한 경로가 없습니다.") 
            
    def log_result(self, msg): 
        self.result_text.insert("end", msg + "\n") 
        self.result_text.see("end")
    
    def log_company_status(self, company_name, status): 
        """사업장명과 처리 상태만 간단히 출력 및 K2B 상태 레이블 업데이트""" 
        self.result_text.insert("end", f"사업장명: {company_name} | 처리상태: {status}\n") 
        self.result_text.see("end")
        
        # K2B 상태 레이블 업데이트
        # 업체명으로 해당 인덱스 찾기
        for i, entry in enumerate(self.entries):
            if entry.get().strip() == company_name:
                if i < len(self.k2b_status_labels):
                    # 성공 상태: "업로드 완료", "정상처리"
                    if status in ["업로드 완료", "정상처리"]:
                        self.k2b_status_labels[i].configure(text="성공", fg_color="green", text_color="white")
                    # 실패 상태: 그 외 모든 상태
                    else:
                        self.k2b_status_labels[i].configure(text="실패", fg_color="red", text_color="white")

                break
    
    def send_email(self, to_email, company_name, year, semester, pdf_path=None, fee_pdf_path=None):
        """
        네이버 SMTP를 사용하여 메일을 발송합니다.
        
        Args:
            to_email: 수신자 이메일 주소
            company_name: 업체명
            year: 년도 (문자열)
            semester: 반기 ("상반기" 또는 "하반기")
            pdf_path: 첨부할 PDF 파일 경로 (필수)
            fee_pdf_path: 첨부할 수수료 내역서 PDF 파일 경로 (선택, 없으면 None)
        
        Returns:
            성공 시 True, 실패 시 False
        """
        try:
            # 네이버 메일 계정 정보 (ID는 고정값, PW는 GUI에서 입력받음)
            naver_id = NAVER_EMAIL_ID
            naver_pw = self.naver_PW_entry.get().strip()
            
            if not naver_pw:
                self.log_result(f"[메일 발송 실패] {company_name}: 네이버 메일 비밀번호가 입력되지 않았습니다.")
                self.log_result(f"[안내] GUI의 '네이버 PW' 입력란에 애플리케이션 비밀번호를 입력하세요.")
                return False
            
            if not to_email:
                self.log_result(f"[메일 발송 실패] {company_name}: 수신자 이메일 주소가 없습니다.")
                return False
            
            # 첨부 파일 필수 확인
            if not pdf_path or not os.path.exists(pdf_path):
                self.log_result(f"[메일 발송 실패] {company_name}: 첨부 파일이 없습니다.")
                self.log_result(f"[파일 경로] {pdf_path if pdf_path else '파일 경로 없음'}")
                self.log_result(f"[안내] 보고서 PDF 파일이 없으면 메일을 발송하지 않습니다.")
                return False
            
            # 네이버 아이디 처리: @naver.com이 포함되어 있으면 제거
            if '@naver.com' in naver_id.lower():
                naver_id = naver_id.lower().replace('@naver.com', '').strip()
            
            # SMTP 서버 설정 (네이버 메일: 포트 465, SSL 필요)
            smtp_server = "smtp.naver.com"
            smtp_port = 465
            
            # 메일 제목 생성: {업체명}-{년도}년 {반기} 작업환경측정결과 보고서 송부
            subject = f"{company_name}-{year}년 {semester} 작업환경측정결과 보고서 송부"
            
            # 메일 본문 생성 (HTML 형식)
            body_html = f"""
<html>
<body style="font-family: 맑은 고딕, Malgun Gothic, sans-serif; font-size: 14px; line-height: 1.6;">
    <p>안녕하십니까!</p>
    
    <p>{year}년 {semester} 작업환경측정결과 보고서 첨부와 같이 송부드리며, 서면은 우편으로 발송 예정이오니 참고하시기 바랍니다.</p>
    
    <p>감사합니다.</p>
    
    <br>
    <hr style="border: none; border-top: 1px solid #cccccc;">
    <p style="font-size: 12px; color: #666666;">
        본 메일 계정은 주식회사 한결작업환경컨설팅의 작업환경측정결과 보고서 발송 전용 계정으로 수신이 불가능합니다. <br>
        회신이나 문의가 필요할 경우 
        <a href="mailto:5678882@naver.com" style="color: #0066cc; font-weight: bold; text-decoration: none;">5678882@naver.com</a>을 이용해 주시기 바랍니다.
    </p>
</body>
</html>
"""
            
            # 메일 내용 설정
            msg = MIMEMultipart()
            # 발신자 이름 설정: "표시 이름 <이메일주소>" 형식
            sender_name = "한결작업환경컨설팅 보고서 발송 전용 계정"
            msg['From'] = f"{Header(sender_name, 'utf-8').encode()} <{naver_id}@naver.com>"
            msg['To'] = to_email
            msg['Subject'] = subject
            msg.attach(MIMEText(body_html, 'html', 'utf-8'))
            
            # PDF 파일 첨부 (보고서)
            try:
                with open(pdf_path, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    
                    # 한글 파일명 인코딩 처리
                    filename = os.path.basename(pdf_path)
                    
                    # RFC 2231 형식으로 인코딩 (한글 파일명 지원)
                    encoded_filename = quote(filename.encode('utf-8'))
                    
                    # RFC 2231 형식 사용 (filename*=UTF-8''encoded)
                    # 이 형식이 한글 파일명을 가장 잘 지원함
                    disposition_str = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
                    part.add_header('Content-Disposition', disposition_str)
                    
                    # 디버깅: 파일명 확인
                    print(f"[DEBUG] 원본 파일명: {filename}")
                    print(f"[DEBUG] 인코딩된 파일명 (RFC 2231): {encoded_filename}")
                    msg.attach(part)
            except Exception as e:
                self.log_result(f"[메일 발송 실패] {company_name}: PDF 첨부 실패 ({e})")
                return False
            
            # 수수료 내역서 PDF 파일 첨부 (선택사항, 있으면 첨부)
            if fee_pdf_path and os.path.exists(fee_pdf_path):
                try:
                    with open(fee_pdf_path, "rb") as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        
                        # 한글 파일명 인코딩 처리
                        filename = os.path.basename(fee_pdf_path)
                        
                        # RFC 2231 형식으로 인코딩 (한글 파일명 지원)
                        encoded_filename = quote(filename.encode('utf-8'))
                        
                        # RFC 2231 형식 사용 (filename*=UTF-8''encoded)
                        disposition_str = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
                        part.add_header('Content-Disposition', disposition_str)
                        
                        # 디버깅: 파일명 확인
                        print(f"[DEBUG] 수수료 내역서 첨부: {filename}")
                        msg.attach(part)
                        self.log_result(f"[메일 발송] {company_name}: 수수료 내역서 첨부됨")
                except Exception as e:
                    # 수수료 내역서 첨부 실패해도 오류로 처리하지 않음 (선택사항이므로)
                    self.log_result(f"[메일 발송] {company_name}: 수수료 내역서 첨부 실패 (무시됨: {e})")
            
            # SMTP 연결 및 메일 발송 (포트 465: SSL 직접 연결)
            self.log_result(f"[메일 발송 진행] {company_name}: SMTP 서버 연결 중 (SSL)...")
            print(f"[DEBUG] 네이버 SMTP 연결 시도: {smtp_server}:{smtp_port} (SSL)")
            print(f"[DEBUG] 사용 ID: {naver_id} (원본: {NAVER_EMAIL_ID})")
            
            # 포트 465는 SSL 직접 연결 (SMTP_SSL 사용, starttls 불필요)
            server = smtplib.SMTP_SSL(smtp_server, smtp_port)
            server.set_debuglevel(1)  # SMTP 디버깅 활성화
            
            self.log_result(f"[메일 발송 진행] {company_name}: 로그인 시도...")
            server.login(naver_id, naver_pw)
            
            self.log_result(f"[메일 발송 진행] {company_name}: 메일 발송 중...")
            text = msg.as_string()
            server.sendmail(msg['From'], msg['To'], text)
            server.quit()
            
            self.log_result(f"[메일 발송 성공] {company_name} → {to_email}")
            return True
            
        except smtplib.SMTPAuthenticationError as e:
            error_msg = str(e)
            error_code = getattr(e, 'smtp_code', 'N/A')
            error_string = getattr(e, 'smtp_error', 'N/A')
            
            self.log_result(f"[메일 발송 실패] {company_name}: 네이버 메일 인증 실패 (오류 코드: {error_code})")
            self.log_result(f"[오류 메시지] Username and Password not accepted")
            self.log_result(f"[사용된 ID] {naver_id} (원본: {NAVER_EMAIL_ID})")
            self.log_result(f"[사용된 포트] 465 (SSL)")
            self.log_result(f"")
            self.log_result(f"=== 네이버 메일 인증 실패 해결 방법 ===")
            self.log_result(f"")
            self.log_result(f"【1단계】POP3/SMTP 설정 확인")
            self.log_result(f"  1. 네이버 메일 접속: https://mail.naver.com")
            self.log_result(f"  2. 우측 상단 설정(톱니바퀴 아이콘) 클릭")
            self.log_result(f"  3. '환경설정' 메뉴 선택")
            self.log_result(f"  4. 왼쪽 메뉴에서 'POP3/SMTP 설정' 클릭")
            self.log_result(f"  5. 'POP3/SMTP 사용' 체크박스가 체크되어 있는지 확인")
            self.log_result(f"  6. 체크되어 있지 않다면 체크 후 '저장' 버튼 클릭")
            self.log_result(f"")
            self.log_result(f"【2단계】애플리케이션 비밀번호 확인 및 생성")
            self.log_result(f"  ⚠️ 중요: 네이버 SMTP 인증은 일반 비밀번호가 아닌 '애플리케이션 비밀번호'를 사용해야 합니다!")
            self.log_result(f"  ※ 2025년 6월 24일부터 POP3/IMAP/SMTP 비밀번호 정책이 변경되어 필수입니다.")
            self.log_result(f"  ")
            self.log_result(f"  [애플리케이션 비밀번호 생성 방법]")
            self.log_result(f"  1. 네이버ID 접속: https://nid.naver.com")
            self.log_result(f"  2. 네이버ID > 보안설정 > 기본보안설정 > 2단계 인증 > 관리")
            self.log_result(f"  3. 네이버 로그인 비밀번호 재확인")
            self.log_result(f"  4. '애플리케이션 비밀번호 관리' 기능 확인")
            self.log_result(f"  5. 사용하려는 애플리케이션 종류 선택 또는 직접 입력 (예: 'SMTP 메일 발송')")
            self.log_result(f"  6. '생성하기' 버튼 클릭")
            self.log_result(f"  7. ⚠️ 생성된 비밀번호를 반드시 복사! (한 번 생성 후 다시 확인 불가)")
            self.log_result(f"  8. 코드의 NAVER_EMAIL_PW 값을 생성된 비밀번호로 변경")
            self.log_result(f"  ")
            self.log_result(f"  [현재 설정]")
            masked_pw = naver_pw[:2] + "*" * (len(naver_pw) - 2) if len(naver_pw) > 2 else "***"
            self.log_result(f"  - 현재 입력된 비밀번호: {masked_pw} (길이: {len(naver_pw)})")
            self.log_result(f"  - 애플리케이션 비밀번호로 변경이 필요합니다!")
            self.log_result(f"")
            self.log_result(f"【3단계】설정 적용")
            self.log_result(f"  - 애플리케이션 비밀번호를 코드에 입력한 후 프로그램을 재시작하세요")
            self.log_result(f"  - 비밀번호를 잃어버린 경우 네이버ID에서 삭제 후 새로 생성하세요")
            
            print(f"[DEBUG] SMTP 인증 오류 상세:")
            print(f"  - 코드: {error_code}")
            print(f"  - 메시지: {error_string}")
            print(f"  - 전체: {error_msg}")
            print(f"  - 사용 ID: {naver_id}")
            print(f"  - 원본 ID: {NAVER_EMAIL_ID}")
            return False
        except smtplib.SMTPException as e:
            error_msg = str(e)
            self.log_result(f"[메일 발송 실패] {company_name}: SMTP 오류 ({error_msg})")
            print(f"[DEBUG] SMTP 오류 상세: {error_msg}")
            return False
        except Exception as e:
            error_msg = str(e)
            self.log_result(f"[메일 발송 실패] {company_name}: 예상치 못한 오류 ({error_msg})")
            print(f"[DEBUG] 예상치 못한 오류 상세: {error_msg}")
            import traceback
            traceback.print_exc()
            return False
    
    def send_emails_manual(self):
        """수동 메일 발송 (버튼 클릭 시)"""
        self.save_config()
        
        year = self.year_entry.get().strip()
        semester = self.semester.get().strip()
        if not year.isdigit():
            messagebox.showerror("입력 오류", "년도는 숫자로 입력해야 합니다.")
            return
        
        base_path = fr"Z:\data\측정팀\측정보고서\{year}년\{semester}"
        selections = [self.entries[i].get().strip() for i in range(10) if self.entries[i].get().strip()]
        
        if not selections:
            messagebox.showwarning("입력 없음", "업체명을 입력하지 않았습니다.")
            return
        
        self.log_result("=" * 50)
        self.log_result("메일 발송 시작")
        self.log_result("=" * 50)
        
        success_count = 0
        fail_count = 0
        
        for i, company_name in enumerate(selections):
            # 실제 폴더명 찾기 (부분 문자열 매칭 + 정규화 비교 사용)
            try:
                all_folders = [f for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
                
                # 정규화 함수 (공백, 괄호 제거)
                def normalize_name(name):
                    return re.sub(r'[()（）\s]', '', name).lower()
                
                # 부분 문자열 매칭 + 정규화 비교
                normalized_input = normalize_name(company_name)
                matching_folders = []
                for folder in all_folders:
                    # 부분 문자열 매칭
                    if company_name in folder or folder in company_name:
                        matching_folders.append(folder)
                    # 정규화된 이름으로 비교
                    elif normalized_input and normalized_input in normalize_name(folder):
                        matching_folders.append(folder)
                    elif normalize_name(folder) and normalize_name(folder) in normalized_input:
                        matching_folders.append(folder)
                
                if not matching_folders:
                    self.log_result(f"[메일 발송 실패] {company_name}: 폴더 없음")
                    self.log_result(f"[디버깅] 검색 경로: {base_path}")
                    self.log_result(f"[디버깅] 입력된 업체명: '{company_name}'")
                    self.log_result(f"[디버깅] 해당 경로의 폴더 목록: {', '.join(all_folders[:10])}...")
                    fail_count += 1
                    # 상태 레이블 업데이트: 실패 (적색)
                    if i < len(self.email_status_labels):
                        self.email_status_labels[i].configure(text="실패", fg_color="red", text_color="white")
                    continue
                
                # 첫 번째 매칭 폴더 사용
                actual_folder_name = matching_folders[0]
                if len(matching_folders) > 1:
                    self.log_result(f"[경고] {company_name}: 여러 폴더가 매칭됨. 첫 번째 폴더 사용: {actual_folder_name}")
                
                target_folder = os.path.join(base_path, actual_folder_name)
                print(f"[DEBUG] 업체명: {company_name} → 실제 폴더: {actual_folder_name}")
                
            except Exception as e:
                self.log_result(f"[메일 발송 실패] {company_name}: 폴더 검색 오류 ({e})")
                fail_count += 1
                # 상태 레이블 업데이트: 실패 (적색)
                if i < len(self.email_status_labels):
                    self.email_status_labels[i].configure(text="실패", fg_color="red", text_color="white")
                continue
            
            if not os.path.exists(target_folder):
                self.log_result(f"[메일 발송 실패] {company_name}: 폴더 경로 확인 실패")
                self.log_result(f"[디버깅] 폴더 경로: {target_folder}")
                fail_count += 1
                # 상태 레이블 업데이트: 실패 (적색)
                if i < len(self.email_status_labels):
                    self.email_status_labels[i].configure(text="실패", fg_color="red", text_color="white")
                continue
            
            # 이메일 주소 가져오기
            email = self.email_labels[i].cget("text").strip()
            if not email:
                # 이메일이 표시되지 않은 경우 엑셀에서 다시 검색
                email = self.find_email_from_excel(year, semester, company_name)
            
            # PDF 파일 찾기: *(보고서-{년도 뒤 2자리}{반기값}).pdf
            # 예: *(보고서-25하).pdf
            year_short = year[-2:]  # 년도 뒤 2자리
            semester_short = '상' if semester == '상반기' else '하'  # 반기 첫글자
            pdf_pattern = f"*(보고서-{year_short}{semester_short}).pdf"
            pdf_files = glob.glob(os.path.join(target_folder, pdf_pattern))
            pdf_path = pdf_files[0] if pdf_files else None
            
            # 수수료 내역서 PDF 파일 찾기: *수수료 내역서({년도 뒤 2자리}{반기값}).pdf
            # 예: *수수료 내역서(25하).pdf (2025년 하반기인 경우만)
            fee_pdf_pattern = f"*수수료 내역서({year_short}{semester_short}).pdf"
            fee_pdf_files = glob.glob(os.path.join(target_folder, fee_pdf_pattern))
            fee_pdf_path = fee_pdf_files[0] if fee_pdf_files else None
            
            # 메일 발송
            if self.send_email(email, company_name, year, semester, pdf_path, fee_pdf_path):
                success_count += 1
                # 상태 레이블 업데이트: 성공 (녹색)
                if i < len(self.email_status_labels):
                    self.email_status_labels[i].configure(text="성공", fg_color="green", text_color="white")
            else:
                fail_count += 1
                # 상태 레이블 업데이트: 실패 (적색)
                if i < len(self.email_status_labels):
                    self.email_status_labels[i].configure(text="실패", fg_color="red", text_color="white")
        
        self.log_result("=" * 50)
        self.log_result(f"메일 발송 완료: 성공 {success_count}건, 실패 {fail_count}건")
        self.log_result("=" * 50)
    
    def send_emails_auto(self, selections, year, semester):
        """자동 메일 발송 (K2B 전송 완료 후)"""
        naver_id = self.naver_ID_entry.get().strip()
        naver_pw = self.naver_PW_entry.get().strip()
        
        if not naver_id or not naver_pw:
            self.log_result("[메일 자동 발송] 네이버 메일 계정 정보가 입력되지 않아 메일을 발송하지 않습니다.")
            return
        
        self.log_result("")
        self.log_result("=" * 50)
        self.log_result("메일 자동 발송 시작")
        self.log_result("=" * 50)
        
        base_path = fr"Z:\data\측정팀\측정보고서\{year}년\{semester}"
        success_count = 0
        fail_count = 0
        
        for i, company_name in enumerate(selections):
            # 실제 폴더명 찾기 (부분 문자열 매칭 + 정규화 비교 사용)
            try:
                all_folders = [f for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
                
                # 정규화 함수 (공백, 괄호 제거)
                def normalize_name(name):
                    return re.sub(r'[()（）\s]', '', name).lower()
                
                # 부분 문자열 매칭 + 정규화 비교
                normalized_input = normalize_name(company_name)
                matching_folders = []
                for folder in all_folders:
                    # 부분 문자열 매칭
                    if company_name in folder or folder in company_name:
                        matching_folders.append(folder)
                    # 정규화된 이름으로 비교
                    elif normalized_input and normalized_input in normalize_name(folder):
                        matching_folders.append(folder)
                    elif normalize_name(folder) and normalize_name(folder) in normalized_input:
                        matching_folders.append(folder)
                
                if not matching_folders:
                    self.log_result(f"[메일 자동 발송 실패] {company_name}: 폴더 없음")
                    # 상태 레이블 업데이트: 실패 (적색)
                    if i < len(self.email_status_labels):
                        self.email_status_labels[i].configure(text="실패", fg_color="red", text_color="white")
                    continue
                
                # 첫 번째 매칭 폴더 사용
                actual_folder_name = matching_folders[0]
                target_folder = os.path.join(base_path, actual_folder_name)
                print(f"[DEBUG] 업체명: {company_name} → 실제 폴더: {actual_folder_name}")
                
            except Exception as e:
                self.log_result(f"[메일 자동 발송 실패] {company_name}: 폴더 검색 오류 ({e})")
                # 상태 레이블 업데이트: 실패 (적색)
                if i < len(self.email_status_labels):
                    self.email_status_labels[i].configure(text="실패", fg_color="red", text_color="white")
                continue
            
            if not os.path.exists(target_folder):
                self.log_result(f"[메일 자동 발송 실패] {company_name}: 폴더 경로 확인 실패")
                # 상태 레이블 업데이트: 실패 (적색)
                if i < len(self.email_status_labels):
                    self.email_status_labels[i].configure(text="실패", fg_color="red", text_color="white")
                continue
            
            # 이메일 주소 가져오기
            email = self.email_labels[i].cget("text").strip()
            if not email:
                # 이메일이 표시되지 않은 경우 엑셀에서 다시 검색
                email = self.find_email_from_excel(year, semester, company_name)
            
            # PDF 파일 찾기: *(보고서-{년도 뒤 2자리}{반기값}).pdf
            # 예: *(보고서-25하).pdf
            year_short = year[-2:]  # 년도 뒤 2자리
            semester_short = '상' if semester == '상반기' else '하'  # 반기 첫글자
            pdf_pattern = f"*(보고서-{year_short}{semester_short}).pdf"
            pdf_files = glob.glob(os.path.join(target_folder, pdf_pattern))
            pdf_path = pdf_files[0] if pdf_files else None
            
            # 수수료 내역서 PDF 파일 찾기: *수수료 내역서({년도 뒤 2자리}{반기값}).pdf
            # 예: *수수료 내역서(25하).pdf (2025년 하반기인 경우만)
            fee_pdf_pattern = f"*수수료 내역서({year_short}{semester_short}).pdf"
            fee_pdf_files = glob.glob(os.path.join(target_folder, fee_pdf_pattern))
            fee_pdf_path = fee_pdf_files[0] if fee_pdf_files else None
            
            # 메일 발송
            if self.send_email(email, company_name, year, semester, pdf_path, fee_pdf_path):
                success_count += 1
                # 상태 레이블 업데이트: 성공 (녹색)
                if i < len(self.email_status_labels):
                    self.email_status_labels[i].configure(text="성공", fg_color="green", text_color="white")
            else:
                fail_count += 1
                # 상태 레이블 업데이트: 실패 (적색)
                if i < len(self.email_status_labels):
                    self.email_status_labels[i].configure(text="실패", fg_color="red", text_color="white")

            self.log_result("=" * 50)
            self.log_result(f"메일 자동 발송 완료: 성공 {success_count}건, 실패 {fail_count}건")
            self.log_result("=" * 50)
    
    def extract_company_status(self, driver):
        """HTML에서 사업장명과 처리 상태를 추출"""
        try:
            # 파일 리스트 그리드에서 정보 추출
            grid_element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '#mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_grid_fileList_bodyGridBandContainerElement_inner'))
            )
            
            # 그리드 내의 행들에서 사업장명과 상태 추출
            rows = grid_element.find_elements(By.CSS_SELECTOR, 'tr')
            for row in rows:
                cells = row.find_elements(By.CSS_SELECTOR, 'td')
                if len(cells) >= 2:  # 사업장명과 상태 컬럼이 있다고 가정
                    company_name = cells[0].text.strip() if cells[0].text else "알 수 없음"
                    status = cells[1].text.strip() if cells[1].text else "알 수 없음"
                    return company_name, status
            
            return "정보 없음", "정보 없음"
        except Exception as e:
            return "추출 실패", f"오류: {str(e)}"
    
    def extract_and_display_all_results(self, driver):
        """그리드에서 사업장명과 처리상태를 추출하여 표시"""
        try:
            # 그리드 컨테이너 요소 확인
            grid_container = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '#mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_grid_fileList_bodyGridBandContainerElement'))
            )
            
            # 모든 그리드 행을 찾기 (gridrow_0, gridrow_1, gridrow_2, ...)
            row_index = 0
            found_any_data = False
            
            while True:
                try:
                    # 사업장명 셀렉터 (cell_0_1, cell_1_1, cell_2_1, ...)
                    company_selector = f'#mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_grid_fileList_body_gridrow_{row_index}_cell_{row_index}_1GridCellTextContainerElement > div'
                    company_element = WebDriverWait(driver, 2).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, company_selector))
                    )
                    
                    # 처리상태 셀렉터 (cell_0_2, cell_1_2, cell_2_2, ...)
                    status_selector = f'#mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_grid_fileList_body_gridrow_{row_index}_cell_{row_index}_2GridCellTextContainerElement > div'
                    status_element = WebDriverWait(driver, 2).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, status_selector))
                    )
                    
                    # 텍스트 추출
                    company_name = company_element.text.strip() if company_element.text else "알 수 없음"
                    status = status_element.text.strip() if status_element.text else "알 수 없음"
                    
                    # 빈 값이 아닌 경우에만 출력
                    if company_name != "알 수 없음" and company_name != "":
                        self.log_company_status(company_name, status)
                        found_any_data = True
                    
                    row_index += 1
                    
                except TimeoutException:
                    # 더 이상 행이 없으면 루프 종료
                    break
                except Exception as e:
                    # 개별 행 처리 중 오류 발생 시 다음 행으로
                    row_index += 1
                    continue
            
            if not found_any_data:
                self.log_result("처리된 업체 정보가 없습니다.")
            
        except TimeoutException:
            self.log_result("그리드 컨테이너를 찾을 수 없습니다. 페이지가 완전히 로드되지 않았을 수 있습니다.")
        except Exception as e:
            self.log_result(f"결과 추출 중 오류 발생: {str(e)}") 

    def reset_all(self): 
        for i in range(10): 
            self.entries[i].delete(0, "end") 
            self.path_labels[i].configure(text="") 
            self.email_labels[i].configure(text="")  # 메일 주소도 초기화
            # K2B 전송 상태도 초기화 (회색, "대기")
            if i < len(self.k2b_status_labels):
                self.k2b_status_labels[i].configure(text="대기", fg_color="#E0E0E0", text_color="black")
            # 메일 발송 상태도 초기화 (회색, "대기")
            if i < len(self.email_status_labels):
                self.email_status_labels[i].configure(text="대기", fg_color="#E0E0E0", text_color="black")

        self.destroy_listbox() 
        self.result_text.delete("1.0", "end") 
        self.save_config() 

    def print_pdf(self, file_path): 
        """ 
        PDFtoPrinter.exe를 사용하여 PDF 파일을 인쇄합니다. 
        가로 방향 페이지(회전이 0인)를 90도 회전하여 인쇄합니다. 
        """ 
        # PyMuPDF가 사용 가능한지 확인
        if not FITZ_AVAILABLE:
            self.log_result("[오류] PyMuPDF가 설치되지 않았습니다. PDF 인쇄 기능을 사용할 수 없습니다.")
            return "PyMuPDF가 설치되지 않았습니다. 'pip install PyMuPDF'를 실행하세요."
            
        temp_dir = os.path.join(os.getcwd(), "temp_prints") 
        if not os.path.exists(temp_dir): 
            os.makedirs(temp_dir) 
            
        rotated_doc = None 
        doc = None 
        rotated_doc_path = None 
        file_to_print = file_path # 기본적으로 원본 파일 사용 

        try: 
            doc = fitz.open(file_path) 
            
            # doc 내에 회전이 필요한 페이지가 있는지 확인 
            needs_rotation = any(page.rect.width > page.rect.height and page.rotation == 0 for page in doc) 

            if needs_rotation: 
                rotated_doc = fitz.open() # 임시 PDF 문서 생성 (Document 객체) 

                for i in range(len(doc)): 
                    page = doc[i] 
                    
                    # 1. 원본 문서에서 페이지 i를 임시 문서의 끝에 복사합니다. 
                    # insert_pdf는 0-based index를 사용합니다. 
                    # from_page와 to_page를 모두 i로 설정하여 단일 페이지를 복사합니다. 
                    rotated_doc.insert_pdf(doc, from_page=i, to_page=i) 
                    
                    # 2. 방금 추가된 페이지(rotated_doc의 마지막 페이지)를 가져옵니다. 
                    rotated_page = rotated_doc[-1] 

                    # 3. 페이지의 가로가 세로보다 길고 (가로 방향) 회전이 적용되지 않았을 경우 
                    if page.rect.width > page.rect.height and page.rotation == 0: 
                        # 페이지의 회전 속성을 -90도로 설정합니다. 
                        rotated_page.set_rotation(-90)  
                        self.log_result(f"[자동] 페이지 {i+1} 가로 감지 → -90도 회전 적용.") 
                    else: 
                        pass # 세로 페이지는 그대로 둠 
                
                # 임시 파일로 저장 
                rotated_doc_path = os.path.join(temp_dir, f"rotated_{os.path.basename(file_path)}") 
                rotated_doc.save(rotated_doc_path, garbage=4) 
                file_to_print = rotated_doc_path 
                self.log_result(f"[자동] 가로 페이지를 회전한 임시 PDF 파일 생성: {os.path.basename(rotated_doc_path)}") 
            
            # PDFtoPrinter.exe를 사용하여 인쇄 
            if PRINTER_NAME: 
                subprocess.run([PDF_TO_PRINTER, file_to_print, PRINTER_NAME], check=True) 
            else: 
                subprocess.run([PDF_TO_PRINTER, file_to_print], check=True) 

            return True 

        except Exception as e: 
            # PyMuPDF 오류를 포함하여 모든 예외를 반환 
            return str(e) 
        finally: 
            # 임시 파일 정리 
            if rotated_doc: 
                rotated_doc.close() 
            if doc: 
                doc.close() 
            
            # 임시 파일 삭제 
            if rotated_doc_path and os.path.exists(rotated_doc_path): 
                time.sleep(1) # 인쇄 스풀링 대기 
                try: 
                    os.remove(rotated_doc_path) 
                except Exception as e: 
                    # 삭제 실패 시 경고 메시지만 출력 
                    self.log_result(f"[경고] 임시 파일 삭제 실패 ({os.path.basename(rotated_doc_path)}): {e}") 

    def on_print(self): 
        win32api.keybd_event(win32con.VK_HANGUL, 0, 0, 0) 
        self.destroy_listbox() 
        year = self.year_entry.get().strip() 
        semester = self.semester.get().strip() 
        if not year.isdigit(): 
            messagebox.showerror("입력 오류", "년도는 숫자로 입력해야 합니다.") 
            return 
        base_path = fr"Z:\data\측정팀\측정보고서\{year}년\{semester}" 
        selections = [self.entries[i].get().strip() for i in range(10) if self.entries[i].get().strip()] 
        if not selections: 
            messagebox.showwarning("입력 없음", "업체명을 입력하지 않았습니다.") 
            return 
        
        for company_name in selections: 
            target_folder = os.path.join(base_path, company_name) 
            if not os.path.exists(target_folder): 
                self.log_result(f"[실패] {company_name} → 폴더 없음") 
                continue 
            
            pdf_files = glob.glob(os.path.join(target_folder, f"*보고서-{year[-2:]}{'상' if semester=='상반기' else '하'}*.pdf")) 
            if not pdf_files: 
                self.log_result(f"[실패] {company_name} → 해당 보고서 파일 없음") 
                continue 
            
            for file_path in pdf_files: 
                result = self.print_pdf(file_path) 
                if result is True: 
                    self.log_result(f"[성공] {company_name} → {os.path.basename(file_path)} 인쇄 완료") 
                else: 
                    self.log_result(f"[실패] {company_name} → {os.path.basename(file_path)} 출력 실패 ({result})") 

    def connect_to_k2b(self): 
        self.save_config() 
        current_ID = self.k2b_ID_entry.get() 
        current_PW = self.k2b_PW_entry.get() 

        if not current_ID or not current_PW: 
            messagebox.showwarning("로그인 정보", "K2B ID와 PW를 입력해주세요.") 
            return 

        driver = None 
        try: 
            options = webdriver.ChromeOptions() 
            options.add_argument("--start-maximized") 
            options.add_experimental_option("detach", True) # 브라우저 꺼짐 방지 옵션
            # Webdriver Manager를 사용하여 드라이버 자동 설치 및 서비스 생성 
            driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options) 

            k2b_url = "https://k2b.kosha.or.kr/index.do" 
            driver.get(k2b_url) 
            time.sleep(2) 

            # Step 1: 로그인 화면 팝업 닫기 (2개) 
            popup_selectors = [ 
                "div#mainframe_VFrameSet_LoginFrame_form_div_popup_361_btn_close", 
                "div#mainframe_VFrameSet_LoginFrame_form_div_popup_360_btn_close" 
            ] 
            for i, selector in enumerate(popup_selectors): 
                try: 
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector))).click() 
                    time.sleep(1) 
                except TimeoutException: 
                    pass 

            # 로그인 페이지 이동 
            time.sleep(2) 
                        
            # Step 2: 로그인 정보 입력 및 클릭 
            # ID 입력 
            ID_input = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mainframe_VFrameSet_LoginFrame_form_div_Login_div_box_edt_mber_ID_input"))) 
            ID_input.send_keys(current_ID) 
            
            # PW 입력 
            PW_input = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='mainframe_VFrameSet_LoginFrame_form_div_Login_div_box_edt_password_input']"))) 
            PW_input.click() 
            PW_input.send_keys(current_PW) 
            
            # 로그인 버튼 클릭 
            login_button = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mainframe_VFrameSet_LoginFrame_form_div_Login_div_box_btn_loginTextBoxElement > div"))) 
            login_button.click() 
            
            # 로그인 성공 확인 
            WebDriverWait(driver, 20).until(EC.title_contains("K2B")) 

            # 내부 화면 팝업 닫기 (1개) 
            popup_selectors = [ 
                "div#mainframe_VFrameSet_MainFrame_form_div_popup_363_btn_closeTextBoxElement > div"
            ] 
            for i, selector in enumerate(popup_selectors): 
                try: 
                    WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector))).click() 
                    time.sleep(1)             
                except TimeoutException: 
                    pass 

            # '파일전송(신)' 버튼이 클릭 가능할 때까지 기다림 
            file_transfer_button = WebDriverWait(driver, 5).until( 
                EC.element_to_be_clickable((By.XPATH, "//div[text()='파일전송(신)']")) 
            ) 
            file_transfer_button.click() 
            time.sleep(3)  

            # 업체별 반복 처리 
            year = self.year_entry.get().strip() 
            semester = self.semester.get().strip() 
            base_path = fr"Z:\data\측정팀\측정보고서\{year}년\{semester}" 
            
            selections = [self.entries[i].get().strip() for i in range(10) if self.entries[i].get().strip()] 

            if not selections: 
                self.log_result("GUI에 입력된 업체명이 없습니다.") 
            else: 
                for company_name in selections: 
                    # 각 업체 처리 전에 팝업이 있는지 확인하고 닫는 로직 추가 
                    try: 
                        WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.popup_close_button"))).click() 
                    except TimeoutException: 
                        pass 
                        
                    target_folder = os.path.join(base_path, company_name) 
                    
                    if not os.path.exists(target_folder): 
                        continue 
                    
                    # *.txt 파일명 추출 
                    txt_files = glob.glob(os.path.join(target_folder, "*.txt")) 
                    
                    if not txt_files: 
                        self.log_company_status(company_name, "txt 파일 없음")
                        continue 
                        
                    # 'XML 추가' 버튼 클릭 
                    add_xml_button = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_div_fileUp_btn_AddTextBoxElement"]/div'))) 
                    add_xml_button.click() 
                    time.sleep(1) # 파일 선택 창이 열리는 것을 대기 

                    # 파일의 절대 경로를 클립보드에 복사 
                    file_path = txt_files[0]  
                    pyperclip.copy(file_path) 
                    time.sleep(1) 
                    
                    # 파일 선택 창에 붙여넣기 및 Enter 
                    pyautogui.hotkey('ctrl', 'v') 
                    time.sleep(1) 
                    pyautogui.press('enter') 
                    time.sleep(3) # 파일 업로드 대기 

                    # '위치도 업로드' 버튼 클릭 
                    try: 
                        location_map_button = WebDriverWait(driver, 20).until( 
                            EC.element_to_be_clickable((By.CSS_SELECTOR, '#mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_div_fileUp_grid_upload_body_gridrow_0_cell_0_2gridCellContainerElement')) 
                        ) 
                        location_map_button.click() 
                        time.sleep(3) 
                    except TimeoutException: 
                        continue 
                        
                    # 다중 JPG 파일 업로드 로직 (개선됨)
                    drawing_folder = os.path.join(target_folder, "도면") 
                    if os.path.exists(drawing_folder): 
                        valid_extensions = ('.jpg', '.jpeg', '.png')
                        jpg_files = [f for f in os.listdir(drawing_folder) if f.lower().endswith(valid_extensions)]
                        jpg_files.sort() # 파일명 기준 오름차순 정렬 추가
                        
                        if jpg_files: 
                            print(f"[DEBUG] {company_name}: JPG 파일 {len(jpg_files)}개 발견")
                            
                            # 2. 파일 선택 창 - 주소창으로 이동 (Alt+D) 후 폴더 경로 입력
                            pyautogui.hotkey('alt', 'd')
                            time.sleep(1)
                            pyperclip.copy(drawing_folder)
                            pyautogui.hotkey('ctrl', 'v')
                            pyautogui.press('enter')
                            time.sleep(1.5) # 폴더 이동 대기
                            
                            # 3. 파일명 입력란으로 이동 (Alt+N) 후 파일명들만 입력
                            # 모든 파일명을 큰따옴표로 묶고 공백으로 연결
                            filenames_only = ' '.join(f'"{f}"' for f in jpg_files)
                            pyperclip.copy(filenames_only)
                            
                            pyautogui.hotkey('alt', 'n')
                            time.sleep(0.5)
                            pyautogui.hotkey('ctrl', 'v')
                            time.sleep(0.5)
                            pyautogui.press('enter')
                            
                            print(f"[DEBUG] {company_name}: 파일명 입력 완료 ({len(filenames_only)}자)")
                            time.sleep(3) # 파일 업로드 완료 대기 시간  

                            # 측정 위치도 '적용' 버튼 클릭 로직 
                            try: 
                                apply_button = WebDriverWait(driver, 20).until( 
                                    EC.element_to_be_clickable((By.XPATH, '//*[@id="mainframe_VFrameSet_MainFrame_DHW00211P01_form_div_Btn_btn_Save"]/div[2]')) 
                                ) 
                                apply_button.click() 
                                time.sleep(2) 
                            except TimeoutException: 
                                continue
                            
                            # XML 업로드 버튼 클릭 
                            try: 
                                upload_button = WebDriverWait(driver, 20).until( 
                                    EC.element_to_be_clickable((By.XPATH, '//*[@id="mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_div_fileUp_btn_UploadTextBoxElement"]/div')) 
                                ) 
                                upload_button.click() 
                                time.sleep(3) # 업로드 완료 대기 
                            except TimeoutException: 
                                continue 
                                
                            # XML 등록 확인 팝업 처리 
                            try: 
                                confirm_button = WebDriverWait(driver, 20).until( 
                                    EC.element_to_be_clickable((By.XPATH, '//*[@id="mainframe_VFrameSet_MainFrame_[1] 건을 업로드 하시겠습니까 ?_form_div_Btn_btn_confirmTextBoxElement"]/div')) 
                                ) 
                                confirm_button.click() 
                                time.sleep(3) # 팝업 처리 대기 
                            except TimeoutException: 
                                continue 
                            

                            # XML 등록 확인 팝업 처리 후 2가지 경우로 분기
                            
                            # 1. '동일한' 메시지 확인
                            try:
                                # '동일한' 텍스트가 포함된 textarea 요소 찾기 (파일명 동적 처리)
                                duplicate_message = WebDriverWait(driver, 5).until(
                                    EC.presence_of_element_located((By.XPATH, "//*[contains(@id, '동일한 파일이 존재합니다') and contains(@id, 'form_tea_message_textarea')]"))
                                )
                                
                                # 찾은 요소의 텍스트 내용을 터미널에만 출력 (GUI에는 출력하지 않음)
                                message_text = duplicate_message.text
                                print(f"[DEBUG] '동일한' 메시지 발견: {message_text}")
                                
                                # 1-1. 확인 버튼 클릭
                                confirm_button = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.XPATH, '//div[text()="확인"]'))
                                )
                                confirm_button.click()
                                time.sleep(2)
                                
                                # 1-2. XML 삭제 버튼 클릭
                                delete_button = WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable((By.CSS_SELECTOR, '#mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_div_fileUp_btn_DelTextBoxElement > div'))
                                )
                                delete_button.click()
                                time.sleep(2)
                                
                                # 1-3. 삭제 확인 팝업의 확인 버튼 클릭
                                delete_confirm_button = WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable((By.XPATH, '//div[text()="확인"]'))
                                )
                                delete_confirm_button.click()
                                time.sleep(2)
                                
                                self.log_company_status(company_name, "동일 파일 삭제 완료")
                                continue  # 다음 업체로 진행
                                
                            except TimeoutException:
                                # 2. '업로드' 메시지 확인
                                try:
                                    # '업로드' 텍스트가 포함된 요소 찾기
                                    upload_message = WebDriverWait(driver, 5).until(
                                        EC.presence_of_element_located((By.XPATH, "//*[contains(text(), '업로드')]"))
                                    )
                                    
                                    # 찾은 요소의 텍스트 내용을 터미널에만 출력 (GUI에는 출력하지 않음)
                                    message_text = upload_message.text
                                    print(f"[DEBUG] '업로드' 메시지 발견: {message_text}")
                                    
                                    # '정상 접수처리 안내' 팝업 처리
                                    success_button = WebDriverWait(driver, 10).until(
                                        EC.element_to_be_clickable((By.XPATH, '//div[text()="확인"]'))
                                    )
                                    success_button.click()
                                    time.sleep(2)
                                    self.log_company_status(company_name, "업로드 완료")
                                        
                                except TimeoutException:
                                    self.log_company_status(company_name, "예상된 메시지 없음") 
                            
                            # 업체 처리 완료 (개별 상태 추출 제거)
                        else:
                            # JPG 파일이 없는 경우 - 바로 XML 삭제 진행
                            self.log_company_status(company_name, "JPG 파일 없음")
                            
                            # 1-2. XML 삭제 버튼 클릭
                            delete_button = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, '#mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_div_fileUp_btn_DelTextBoxElement > div'))
                            )
                            delete_button.click()
                            time.sleep(2)
                            
                            # 1-3. 삭제 확인 팝업의 확인 버튼 클릭
                            delete_confirm_button = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, '//div[text()="확인"]'))
                            )
                            delete_confirm_button.click()
                            time.sleep(2)
                            
                            self.log_company_status(company_name, "JPG 파일 없음으로 인한 XML 삭제 완료")
                            continue  # 다음 업체로 진행
                    else:
                        # 도면 폴더가 없는 경우 - 바로 XML 삭제 진행
                        self.log_company_status(company_name, "도면 폴더 없음")
                        
                        # XML 삭제 버튼 클릭
                        delete_button = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, '#mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_div_fileUp_btn_DelTextBoxElement > div'))
                        )
                        delete_button.click()
                        time.sleep(2)
                        
                        # 삭제 확인 팝업의 확인 버튼 클릭
                        delete_confirm_button = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//div[text()="확인"]'))
                        )
                        delete_confirm_button.click()
                        time.sleep(2)
                        
                        self.log_company_status(company_name, "도면 폴더 없음으로 인한 XML 삭제 완료")
                        continue  # 다음 업체로 진행 
                            
        except (NoSuchElementException, TimeoutException) as e: 
            self.log_result(f"자동화 오류: {e}") 
            messagebox.showerror("자동화 실패", f"페이지 요소 오류가 발생했습니다.\n오류: {e}") 
            # 오류 발생 시에만 브라우저 닫기
            if driver: 
                driver.quit()
        except WebDriverException as e: 
            self.log_result(f"WebDriver 오류: {e}") 
            messagebox.showerror("WebDriver 오류", f"웹 드라이버 실행 문제 발생.\n오류: {e}") 
            # 오류 발생 시에만 브라우저 닫기
            if driver: 
                driver.quit()
        except Exception as e: 
            self.log_result(f"예상치 못한 오류: {e}") 
            messagebox.showerror("오류", f"K2B 전송 중 오류가 발생했습니다: {e}") 
            # 오류 발생 시에만 브라우저 닫기
            if driver: 
                driver.quit()
        # 업체 처리 여부와 관계없이 파일 접수 현황에서 결과 추출
        try:
            # 10초 대기 (데이터 처리 시간 확보)
            self.log_result("데이터 처리를 위해 10초 대기 후 조회합니다...")
            time.sleep(10)

            # 조회 버튼 클릭 (현황 갱신)
            try:
                search_btn_selector = '#mainframe_VFrameSet_MainFrame_form_div_Form_div_Work_103017203_div_Work_div_Search_btn_SearchTextBoxElement > div'
                search_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, search_btn_selector))
                )
                search_button.click()
                self.log_result("접수 현황을 갱신했습니다.")
                time.sleep(3) # 조회 결과 로딩 대기
            except Exception as e:
                self.log_result(f"조회 버튼 클릭 실패 (현황이 갱신되지 않았을 수 있습니다): {str(e)}")

            # 2행 빈열 추가
            self.log_result("")
            self.log_result("")
            
            if selections:
                self.log_result("=== 금일 K2B 전송 결과 ===")
            else:
                self.log_result("=== 파일 접수 현황 ===")
            self.extract_and_display_all_results(driver)
        except Exception as e:
            self.log_result(f"결과 추출 실패: {str(e)}")
        
        # K2B 전송 보고 완료 메시지 표시
        self.log_result("")
        self.log_result("=" * 50)
        self.log_result("K2B 전송 보고 완료")
        self.log_result("=" * 50)
        
        # K2B 전송 완료 후 자동 메일 발송 (비활성화 - 메일 발송 버튼으로 분리)
        # if selections:
        #     self.send_emails_auto(selections, year, semester)
        
        # 모든 업체 처리 완료 후 브라우저 닫기
        if selections:
            self.log_result("모든 업체 처리가 완료되었습니다. (브라우저 유지)")
            # if driver:
            #     driver.quit()
        else:
            self.log_result("파일 접수 현황 조회가 완료되었습니다. (브라우저 유지)")
            # if driver:
            #     driver.quit() 

if __name__ == "__main__": 
    app = ReportPrinterApp() 
    app.mainloop()
