import customtkinter as ctk
from tkinter import filedialog, messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import threading
import time
import os
import subprocess
import platform
import re
from openpyxl import load_workbook
import json
import sys
from datetime import datetime


def normalize_contact_phone_suffix(value):
    """공단 화면의 고정 010 선택 상자 뒤에 입력할 가입자 번호만 반환합니다."""
    digits = re.sub(r"\D", "", str(value or ""))
    if digits.startswith("010"):
        digits = digits[3:]
    if len(digits) not in (7, 8):
        raise ValueError("담당자 휴대전화는 010 번호여야 합니다.")
    return digits


def classify_employee_check(value):
    text = str(value or "").strip()
    if not text:
        return "NO_INFO"
    if "50인 이상" in text:
        return "OVER_50"
    if "50인 미만" in text or ("%" in text and "50인 이상" not in text):
        return "OK"
    return "EMPLOYEE_CHECK_FAILED"


def has_application_success_marker(value):
    text = str(value or "")
    return any(marker in text for marker in (
        "신청이 완료", "신청 완료", "접수가 완료", "접수번호",
    ))


# customtkinter 설정
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


class HealthProgramAutomation:
    def get_resource_path(self, relative_path):
        """리소스 파일의 절대 경로를 반환합니다. (PyInstaller 패키징 대응)"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("건강디딤돌 신청 자동화 프로그램")
        self.root.geometry("1920x1080+0+0")  # 좌측 상단에 위치
        
        # 아이콘 설정
        icon_path = self.get_resource_path("건강디딤돌.ico")
        if os.path.exists(icon_path):
            try:
                self.root.iconbitmap(icon_path)
            except Exception:
                pass
        
        # 변수 초기화
        self.excel_file_path = None
        self.sheet_name = '건강디딤돌신청data_DB'
        self.driver = None
        self.is_running = False
        self.is_paused = False
        self.pause_event = threading.Event()  # 일시 중지/재개 제어용
        self.pause_event.set()  # 초기값은 실행 상태
        
        # GUI 구성
        self.setup_gui()
        
        # 설정 불러오기
        self.load_settings()
        
        # 윈도우 종료 프로토콜 연결
        self.root.protocol("WM_DELETE_WINDOW", self.exit_program)
        
    def setup_gui(self):
        """GUI 구성 요소 생성"""
        # 메인 프레임
        main_frame = ctk.CTkFrame(self.root)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # 제목
        title_label = ctk.CTkLabel(main_frame, text="건강디딤돌 신청 자동화 프로그램", 
                                   font=ctk.CTkFont(size=28, weight="bold"))
        title_label.pack(pady=(20, 30))
        
        # 입력 영역 프레임
        input_frame = ctk.CTkFrame(main_frame)
        input_frame.pack(fill="x", padx=20, pady=10)
        
        # 신청년도와 반기를 같은 줄에 배치
        year_half_frame = ctk.CTkFrame(input_frame)
        year_half_frame.pack(fill="x", padx=20, pady=10)
        
        # 신청년도 입력
        ctk.CTkLabel(year_half_frame, text="신청년도:", font=ctk.CTkFont(size=14)).pack(side="left", padx=10)
        self.year_entry = ctk.CTkEntry(year_half_frame, width=150, font=ctk.CTkFont(size=14))
        self.year_entry.pack(side="left", padx=10)
        self.year_entry.insert(0, "2026")
        
        # 반기 선택 (같은 줄에 배치)
        ctk.CTkLabel(year_half_frame, text="반기:", font=ctk.CTkFont(size=14)).pack(side="left", padx=(30, 10))
        self.half_combo = ctk.CTkComboBox(year_half_frame, values=["상반기", "하반기"], width=150, font=ctk.CTkFont(size=14))
        self.half_combo.pack(side="left", padx=10)
        self.half_combo.set("상반기")
        
        # 엑셀 파일 선택
        excel_frame = ctk.CTkFrame(input_frame)
        excel_frame.pack(fill="x", padx=20, pady=10)
        ctk.CTkLabel(excel_frame, text="엑셀 파일:", font=ctk.CTkFont(size=14)).pack(side="left", padx=10)
        self.excel_path_label = ctk.CTkLabel(excel_frame, text="파일을 선택해주세요", 
                                             font=ctk.CTkFont(size=12), text_color="gray")
        self.excel_path_label.pack(side="left", padx=10, fill="x", expand=True)
        select_btn = ctk.CTkButton(excel_frame, text="파일 선택", command=self.select_excel_file, width=100)
        select_btn.pack(side="right", padx=10)
        
        # 버튼 영역 (중앙 정렬)
        button_frame = ctk.CTkFrame(main_frame)
        button_frame.pack(fill="x", padx=20, pady=20)
        
        # 버튼을 중앙에 배치하기 위한 내부 프레임
        button_inner_frame = ctk.CTkFrame(button_frame)
        button_inner_frame.pack(expand=True)
        
        self.apply_btn = ctk.CTkButton(button_inner_frame, text="신청", command=self.run_apply_workflow,
                                       width=200, height=50, font=ctk.CTkFont(size=16, weight="bold"))
        self.apply_btn.pack(side="left", padx=20, pady=10)
        
        self.check_btn = ctk.CTkButton(button_inner_frame, text="결과 확인", command=self.open_excel_file,
                                      width=200, height=50, font=ctk.CTkFont(size=16, weight="bold"))
        self.check_btn.pack(side="left", padx=20, pady=10)
        
        # 제어 버튼 영역 (일시 중지, 다시 시작, 종료)
        control_button_frame = ctk.CTkFrame(main_frame)
        control_button_frame.pack(fill="x", padx=20, pady=10)
        
        # 제어 버튼을 중앙에 배치하기 위한 내부 프레임
        control_inner_frame = ctk.CTkFrame(control_button_frame)
        control_inner_frame.pack(expand=True)
        
        self.pause_btn = ctk.CTkButton(control_inner_frame, text="일시 중지", command=self.pause_workflow,
                                       width=150, height=40, font=ctk.CTkFont(size=14, weight="bold"),
                                       fg_color="#FFA500", hover_color="#FF8C00", text_color="black", text_color_disabled="black")
        self.pause_btn.pack(side="left", padx=10, pady=10)
        self.pause_btn.configure(state="disabled", text_color="black")  # 초기에는 비활성화
        
        self.resume_btn = ctk.CTkButton(control_inner_frame, text="다시 시작", command=self.resume_workflow,
                                       width=150, height=40, font=ctk.CTkFont(size=14, weight="bold"),
                                       fg_color="#32CD32", hover_color="#228B22", text_color="black", text_color_disabled="black")
        self.resume_btn.pack(side="left", padx=10, pady=10)
        self.resume_btn.configure(state="disabled", text_color="black")  # 초기에는 비활성화
        
        self.exit_btn = ctk.CTkButton(control_inner_frame, text="종료", command=self.exit_program,
                                      width=150, height=40, font=ctk.CTkFont(size=14, weight="bold"),
                                      fg_color="#DC143C", hover_color="#B22222", text_color="black")
        self.exit_btn.pack(side="left", padx=10, pady=10)
        
        # 진행 사항 표시 영역
        progress_frame = ctk.CTkFrame(main_frame)
        progress_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # 진행 사항 제목과 로그 초기화 버튼을 같은 줄에 배치
        progress_header_frame = ctk.CTkFrame(progress_frame)
        progress_header_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(progress_header_frame, text="진행 사항", font=ctk.CTkFont(size=16, weight="bold")).pack(side="left")
        
        # 로그 초기화 버튼 (보라색 계열 파스텔톤)
        clear_log_btn = ctk.CTkButton(progress_header_frame, text="로그 초기화", command=self.clear_log,
                                      width=120, height=30, font=ctk.CTkFont(size=12),
                                      fg_color="#B19CD9", hover_color="#9B87C7", text_color="white")
        clear_log_btn.pack(side="right", padx=10)
        
        # 스크롤 가능한 텍스트 영역
        self.progress_text = ctk.CTkTextbox(progress_frame, font=ctk.CTkFont(size=12))
        self.progress_text.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
    def load_settings(self):
        """설정 파일(settings.json) 로드 및 GUI 적용"""
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "settings.json")
        if os.path.exists(config_path):
            try:
                with open(config_path, "r", encoding="utf-8") as f:
                    config = json.load(f)
                
                # 년도 복구
                year = config.get("year", "2026")
                self.year_entry.delete(0, "end")
                self.year_entry.insert(0, year)
                
                # 반기 복구
                half = config.get("half", "상반기")
                if half in ["상반기", "하반기"]:
                    self.half_combo.set(half)
                
                # 엑셀 경로 복구
                excel_path = config.get("excel_file_path")
                if excel_path and os.path.exists(excel_path):
                    self.excel_file_path = excel_path
                    self.excel_path_label.configure(text=os.path.basename(excel_path), text_color="black")
                    self.update_progress(f"설정 파일에서 엑셀 파일 복구: {os.path.basename(excel_path)}")
                
                self.update_progress("이전 설정 정보를 불러왔습니다.")
            except Exception as e:
                self.update_progress(f"설정 파일을 불러오는 중 오류 발생: {str(e)}")

    def save_settings(self):
        """현재 GUI 설정을 파일(settings.json)로 저장"""
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "settings.json")
        try:
            config = {
                "year": self.year_entry.get().strip(),
                "half": self.half_combo.get(),
                "excel_file_path": self.excel_file_path
            }
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=4)
        except Exception as e:
            self.update_progress(f"설정 저장 실패: {str(e)}")

    def select_excel_file(self):
        """엑셀 파일 선택 다이얼로그"""
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_file_path = file_path
            self.excel_path_label.configure(text=os.path.basename(file_path), text_color="black")
            self.update_progress(f"엑셀 파일 선택: {os.path.basename(file_path)}")
            self.save_settings()
    
    def open_excel_file(self):
        """엑셀 파일 열기"""
        if not self.excel_file_path:
            messagebox.showwarning("알림", "엑셀 파일을 먼저 선택해주세요.")
            return
        
        if not os.path.exists(self.excel_file_path):
            messagebox.showwarning("알림", "엑셀 파일을 찾을 수 없습니다.")
            return
        
        try:
            # 운영체제에 따라 적절한 명령어 사용
            system = platform.system()
            if system == "Windows":
                os.startfile(self.excel_file_path)
            elif system == "Darwin":  # macOS
                subprocess.run(["open", self.excel_file_path])
            else:  # Linux
                subprocess.run(["xdg-open", self.excel_file_path])
            self.update_progress(f"엑셀 파일 열기: {os.path.basename(self.excel_file_path)}")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 파일을 열 수 없습니다: {str(e)}")
            self.update_progress(f"엑셀 파일 열기 실패: {str(e)}")
    
    def check_excel_open(self):
        """엑셀 파일이 열려있는지 확인"""
        if not self.excel_file_path:
            return False, "엑셀 파일을 선택해주세요."
        
        if not os.path.exists(self.excel_file_path):
            return False, "엑셀 파일을 찾을 수 없습니다."
        
        try:
            # 파일을 읽기 전용 모드로 열어보기 시도
            workbook = load_workbook(self.excel_file_path, read_only=True)
            workbook.close()
            return True, None
        except PermissionError:
            return False, "엑셀 파일이 열려있습니다. 파일을 닫고 다시 시도해주세요."
        except Exception as e:
            return False, f"엑셀 파일 확인 중 오류: {str(e)}"
    
    def update_progress(self, message):
        """진행 상황 업데이트"""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        self.progress_text.insert("end", f"[{timestamp}] {message}\n")
        self.progress_text.see("end")
        self.root.update_idletasks()
    
    def clear_log(self):
        """로그 초기화"""
        self.progress_text.delete("1.0", "end")
        # 초기화 완료 메시지 추가 (update_progress를 사용하지 않고 직접 추가)
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        self.progress_text.insert("end", f"[{timestamp}] 로그가 초기화되었습니다.\n")
        self.progress_text.see("end")
        self.root.update_idletasks()
    
    def init_webdriver(self):
        """웹 드라이버 초기화"""
        if self.driver:
            return self.driver
        
        try:
            chrome_options = Options()
            chrome_options.add_argument("--start-maximized")
            self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
            return self.driver
        except Exception as e:
            self.update_progress(f"웹 드라이버 초기화 실패: {str(e)}")
            raise
    
    def load_excel_file(self):
        """엑셀 파일 로드"""
        if not self.excel_file_path:
            raise ValueError("엑셀 파일을 선택해주세요.")
        
        if not os.path.exists(self.excel_file_path):
            raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {self.excel_file_path}")
        
        try:
            workbook = load_workbook(self.excel_file_path)
            if self.sheet_name not in workbook.sheetnames:
                workbook.close()
                raise ValueError(f"시트 '{self.sheet_name}'를 찾을 수 없습니다.")
            sheet = workbook[self.sheet_name]
            return workbook, sheet
        except PermissionError:
            raise PermissionError("엑셀 파일이 열려있습니다. 파일을 닫고 다시 시도해주세요.")
        except Exception as e:
            raise Exception(f"엑셀 파일 로드 중 오류: {str(e)}")
    
    def run_apply_workflow(self):
        """신청 버튼 클릭 시 전체 워크플로우 실행"""
        if self.is_running:
            messagebox.showwarning("알림", "이미 작업이 실행 중입니다.")
            return
        
        # 엑셀 파일 열림 확인
        is_ok, error_msg = self.check_excel_open()
        if not is_ok:
            messagebox.showwarning("알림", error_msg)
            return
        
        # 입력값 확인
        if not self.excel_file_path:
            messagebox.showwarning("알림", "엑셀 파일을 선택해주세요.")
            return
        
        year = self.year_entry.get().strip()
        half = self.half_combo.get()
        
        if not year:
            messagebox.showwarning("알림", "신청년도를 입력해주세요.")
            return
        
        # 현재 설정 저장
        self.save_settings()
        
        # 별도 스레드에서 실행
        self.is_running = True
        self.is_paused = False
        self.pause_event.set()  # 실행 상태로 설정
        self.apply_btn.configure(state="disabled")
        self.check_btn.configure(state="disabled")
        self.pause_btn.configure(state="normal", text_color="black")  # 일시 중지 버튼 활성화
        self.resume_btn.configure(state="disabled", text_color="black")  # 다시 시작 버튼 비활성화
        
        thread = threading.Thread(target=self._apply_workflow_thread, args=(year, half), daemon=True)
        thread.start()
    
    def run_check_workflow(self):
        """확인 버튼 클릭 시 결과 확인만 실행"""
        if self.is_running:
            messagebox.showwarning("알림", "이미 작업이 실행 중입니다.")
            return
        
        # 엑셀 파일 열림 확인
        is_ok, error_msg = self.check_excel_open()
        if not is_ok:
            messagebox.showwarning("알림", error_msg)
            return
        
        # 입력값 확인
        if not self.excel_file_path:
            messagebox.showwarning("알림", "엑셀 파일을 선택해주세요.")
            return
        
        year = self.year_entry.get().strip()
        half = self.half_combo.get()
        
        if not year:
            messagebox.showwarning("알림", "신청년도를 입력해주세요.")
            return
        
        # 현재 설정 저장
        self.save_settings()
        
        # 별도 스레드에서 실행
        self.is_running = True
        self.apply_btn.configure(state="disabled")
        self.check_btn.configure(state="disabled")
        
        thread = threading.Thread(target=self._check_workflow_thread, args=(year, half), daemon=True)
        thread.start()
    
    def _apply_workflow_thread(self, year, half):
        """신청 워크플로우 실행 (스레드)"""
        try:
            self.update_progress("=" * 50)
            self.update_progress("신청 작업 시작")
            self.update_progress(f"신청년도: {year}, 반기: {half}")
            
            # 1. 신청 처리
            self.update_progress("\n[1단계] 신청 처리 시작")
            self.apply_health_program(year, half)
            
            # 2. 결과 확인 처리
            self.update_progress("\n[2단계] 결과 확인 처리 시작")
            self.check_result(year, half)
            
            # 3. 엑셀 파일 저장
            self.update_progress("\n[3단계] 엑셀 파일 저장")
            try:
                workbook = load_workbook(self.excel_file_path)
                workbook.save(self.excel_file_path)
                workbook.close()
                self.update_progress("엑셀 파일 저장 완료")
            except Exception as e:
                self.update_progress(f"엑셀 파일 저장 중 오류: {str(e)}")
            
            self.update_progress("\n모든 작업이 완료되었습니다.")
            messagebox.showinfo("완료", "모든 작업이 완료되었습니다.")
            
        except Exception as e:
            self.update_progress(f"\n오류 발생: {str(e)}")
            messagebox.showerror("오류", f"작업 중 오류가 발생했습니다: {str(e)}")
        finally:
            self.is_running = False
            self.is_paused = False
            self.pause_event.set()
            self.apply_btn.configure(state="normal")
            self.check_btn.configure(state="normal")
            self.pause_btn.configure(state="disabled", text_color="black")
            self.resume_btn.configure(state="disabled", text_color="black")
            if self.driver:
                try:
                    self.driver.quit()
                    self.driver = None
                except:
                    pass
    
    def pause_workflow(self):
        """일시 중지"""
        if self.is_running and not self.is_paused:
            self.is_paused = True
            self.pause_event.clear()  # 일시 중지 상태로 설정
            self.pause_btn.configure(state="disabled", text_color="black")
            self.resume_btn.configure(state="normal", text_color="black")
            self.update_progress("\n작업이 일시 중지되었습니다.")
    
    def resume_workflow(self):
        """다시 시작"""
        if self.is_running and self.is_paused:
            self.is_paused = False
            self.pause_event.set()  # 실행 상태로 설정
            self.pause_btn.configure(state="normal", text_color="black")
            self.resume_btn.configure(state="disabled", text_color="black")
            self.update_progress("\n작업을 다시 시작합니다.")
    
    def exit_program(self):
        """프로그램 종료"""
        if self.is_running:
            result = messagebox.askyesno("확인", "작업이 실행 중입니다. 정말 종료하시겠습니까?")
            if not result:
                return
            # 작업 중지
            self.is_running = False
            self.is_paused = False
            self.pause_event.set()
            # 드라이버 종료
            if self.driver:
                try:
                    self.driver.quit()
                    self.driver = None
                except:
                    pass
        
        # 현재 GUI 설정 저장
        self.save_settings()
        
        self.root.quit()
        self.root.destroy()
    
    def _check_workflow_thread(self, year, half):
        """결과 확인 워크플로우 실행 (스레드)"""
        try:
            self.update_progress("=" * 50)
            self.update_progress("결과 확인 작업 시작")
            self.update_progress(f"신청년도: {year}, 반기: {half}")
            
            # 결과 확인 처리만 실행
            self.check_result(year, half)
            
            # 엑셀 파일 저장
            self.update_progress("\n엑셀 파일 저장")
            try:
                workbook = load_workbook(self.excel_file_path)
                workbook.save(self.excel_file_path)
                workbook.close()
                self.update_progress("엑셀 파일 저장 완료")
            except Exception as e:
                self.update_progress(f"엑셀 파일 저장 중 오류: {str(e)}")
            
            self.update_progress("\n결과 확인 작업이 완료되었습니다.")
            messagebox.showinfo("완료", "결과 확인 작업이 완료되었습니다.")
            
        except Exception as e:
            self.update_progress(f"\n오류 발생: {str(e)}")
            messagebox.showerror("오류", f"작업 중 오류가 발생했습니다: {str(e)}")
        finally:
            self.is_running = False
            self.is_paused = False
            self.pause_event.set()
            self.apply_btn.configure(state="normal")
            self.check_btn.configure(state="normal")
            self.pause_btn.configure(state="disabled", text_color="black")
            self.resume_btn.configure(state="disabled", text_color="black")
            if self.driver:
                try:
                    self.driver.quit()
                    self.driver = None
                except:
                    pass
    
    def wait_if_paused(self):
        """일시 중지 상태면 대기"""
        if self.is_paused:
            self.update_progress("일시 중지 중... 다시 시작 버튼을 클릭하세요.")
            self.pause_event.wait()  # 다시 시작될 때까지 대기
    
    def apply_health_program(self, year, half):
        """신청 자동화 로직"""
        workbook = None
        try:
            workbook, sheet = self.load_excel_file()
            self.update_progress("엑셀 파일을 성공적으로 불러왔습니다.")
        except Exception as e:
            self.update_progress(f"엑셀 파일을 불러오는 데 실패했습니다: {str(e)}")
            return
        
        try:
            # A열에 값이 있는 마지막 행 찾기
            last_row_with_data = 1
            for row in range(sheet.max_row, 1, -1):
                if sheet[f'A{row}'].value is not None and str(sheet[f'A{row}'].value).strip() != '':
                    last_row_with_data = row
                    break
            
            self.update_progress(f"A열에 값이 있는 마지막 행: {last_row_with_data}행")
            if last_row_with_data < 2:
                self.update_progress("처리할 데이터가 없습니다.")
                return
            
            # Chrome 드라이버 초기화
            self.init_webdriver()
            self.update_progress("크롬 브라우저를 실행합니다.")
            
            url = "https://portal.kosha.or.kr/business-apply-search/health-support/step-stone/info"
            self.driver.get(url)
            self.update_progress(f"URL: {url} 로 접속합니다.")
            
            # 처리할 행 수 계산
            total_rows = 0
            for row in range(2, last_row_with_data + 1):
                h_value = sheet[f'H{row}'].value
                if h_value != "○":
                    total_rows += 1
            
            self.update_progress(f"처리할 행 수: {total_rows}개")
            
            processed = 0
            # A2부터 A열에 값이 있는 마지막 행까지 반복
            for row in range(2, last_row_with_data + 1):
                # 신청 완료된 행은 건너뛰기
                h_value = sheet[f'H{row}'].value
                if h_value == "○":
                    continue
                
                # 엑셀에서 데이터 읽기
                biz_id = sheet[f'A{row}'].value
                biz_start_no = sheet[f'B{row}'].value
                biz_name = sheet[f'D{row}'].value
                biz_address = sheet[f'E{row}'].value
                biz_contact_name = sheet[f'F{row}'].value
                biz_contact_phone = sheet[f'G{row}'].value
                
                # 모든 데이터가 정상적으로 읽어졌는지 확인
                if not all([biz_id, biz_start_no, biz_contact_name, biz_contact_phone]):
                    self.update_progress(f"{row}행에 누락된 데이터가 있습니다. 해당 행을 건너뜁니다.")
                    continue
                
                # 일시 중지 체크
                self.wait_if_paused()
                if not self.is_running:
                    break
                
                processed += 1
                self.update_progress(f"\n--- [{processed}/{total_rows}] {row}행 데이터로 자동화 시작 ---")
                
                try:
                    # 신청 프로세스 실행
                    result_status = self._process_application(row, biz_id, biz_start_no, biz_name, biz_address, 
                                             biz_contact_name, biz_contact_phone, url, half, year)
                    
                    if result_status == "OK":
                        # 신청 완료된 행에 '○' 기재 (H열: 신청 여부)
                        sheet[f'H{row}'] = "○"
                        workbook.save(self.excel_file_path)
                        self.update_progress(f"{row}행의 신청이 완료되어 H열에 '○'를 기재했습니다.")
                    elif result_status == "OVER_50":
                        sheet[f'H{row}'] = "50인 이상"
                        workbook.save(self.excel_file_path)
                        self.update_progress(f"{row}행: 50인 이상 사업장으로 신청 취소. H열에 '50인 이상' 기재.")
                    elif result_status == "NO_INFO":
                        sheet[f'H{row}'] = "정보없음"
                        workbook.save(self.excel_file_path)
                        self.update_progress(f"{row}행: 근로자 수 정보 없음으로 신청 취소. H열에 '정보없음' 기재.")
                    
                    if result_status != "OK":
                        # 신청 취소 시 다음 처리를 위해 잠시 대기
                        time.sleep(2)

                    
                except Exception as e:
                    self.update_progress(f"--- {row}행 자동화 중 오류 발생: {str(e)} ---")
                    # 오류 발생 시 신청 여부 기록 (H열: 신청 여부)
                    try:
                        sheet[f'H{row}'] = "X"
                        workbook.save(self.excel_file_path)
                        self.update_progress(f"{row}행의 신청 실패를 H열에 'X'로 기록했습니다.")
                    except:
                        pass
                    # 오류 발생 시 해당 행은 건너뛰고 다음 행으로 진행
                    try:
                        self.driver.get(url)
                        time.sleep(3)
                    except:
                        self.update_progress("메인 페이지로 돌아가는 중에도 오류 발생.")
                        break
            
            self.update_progress("\n신청 작업이 완료되었습니다.")
        except Exception as e:
            self.update_progress(f"초기화 중 오류 발생: {str(e)}")
        finally:
            if workbook:
                workbook.close()
    
    def _process_application(self, row, biz_id, biz_start_no, biz_name, biz_address, 
                            biz_contact_name, biz_contact_phone, url, half, year=None):
        """개별 신청 프로세스 처리"""
        # 첫 번째 '신청하기' 버튼 클릭
        apply_button_xpath1 = '//*[@id="contents"]/header/div[2]/button'
        apply_button1 = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, apply_button_xpath1))
        )
        apply_button1.click()
        time.sleep(1.5)
        
        # "50인 미만 사업장 신청하기" 클릭
        apply_link_xpath = '//*[@id="contents"]/div[1]/div/a[1]'
        apply_link = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, apply_link_xpath))
        )
        apply_link.click()
        time.sleep(1.5)

        requested_year = str(year or "").strip()
        if not requested_year or requested_year != str(datetime.now().year):
            self.update_progress("  -> 신청 연도가 현재 접수 연도와 다르므로 자동 신청을 중단합니다.")
            return "YEAR_CHECK_REQUIRED"
        portal_year_text = self.driver.find_element(By.TAG_NAME, "body").text
        if requested_year not in portal_year_text:
            self.update_progress("  -> 공단 신청 화면에서 요청 연도를 확인할 수 없어 중단합니다.")
            return "YEAR_CHECK_REQUIRED"
        
        # "0000년 50인미만 비용지원 비율 등 안내" 팝업창 확인
        popup_button_xpath = '//*[@id="app"]/div[7]/div/section/footer/button[2]/span'
        popup_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, popup_button_xpath))
        )
        popup_button.click()
        time.sleep(2)
        
        # 1단계 약관동의 "모두 동의" 체크
        terms_checkbox_xpath1 = '//*[@id="contents"]/section[1]/div[1]/label'
        terms_checkbox1 = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, terms_checkbox_xpath1))
        )
        terms_checkbox1.click()
        time.sleep(1.5)
        
        # "확인 및 서약" 체크
        terms_checkbox_xpath2 = '//*[@id="contents"]/section[1]/section/div/section/footer/label'
        terms_checkbox2 = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, terms_checkbox_xpath2))
        )
        terms_checkbox2.click()
        time.sleep(1.5)
        
        # "다음" 단계 버튼 클릭
        next_button_xpath = '//*[@id="contents"]/footer/div[2]/button'
        next_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, next_button_xpath))
        )
        next_button.click()
        
        # 2단계 사업장명 '검색' 버튼이 렌더링될 때까지 대기
        search_button_xpath = '//*[@id="contents"]/section/section[1]/div[1]/div/button'
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, search_button_xpath))
        )
        time.sleep(2)
        
        # 2단계 사업장명 '검색' 버튼 클릭
        search_button = self.driver.find_element(By.XPATH, search_button_xpath)
        search_button.click()
        time.sleep(1.5)
        
        # 사업장 관리 번호 입력
        input_field_xpath = '//*[@id="searchText"]'
        input_field = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, input_field_xpath))
        )
        input_field.send_keys(str(biz_id))
        time.sleep(1.5)
        
        # 사업개시번호 입력
        biz_start_no_field = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#bizStrtNo'))
        )
        biz_start_no_field.send_keys(str(biz_start_no))
        time.sleep(1.5)
        
        # 사업장 조회 "조회" 버튼 클릭
        search_button_xpath_final = '/html/body/div[3]/div/section/div/div[1]/button'
        search_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, search_button_xpath_final))
        )
        search_button.click()
        time.sleep(1.5)
        
        # 사업장 목록에서 라디오 버튼 선택
        radio_selector = 'body > div:nth-child(4) > div > section > div > div.table > table > tbody > tr > td.case0 > label > input[type=radio]'
        radio_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, radio_selector))
        )
        self.driver.execute_script("arguments[0].click();", radio_button)
        time.sleep(1.5)
        
        # 사업장 조회 확인버튼 클릭
        confirm_button_xpath = '/html/body/div[3]/div/section/footer/button[2]'
        confirm_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, confirm_button_xpath))
        )
        confirm_button.click()
        time.sleep(1.5)
        
        # 대표자명 입력
        representative_xpath = '//*[@id="businessRvNm"]'
        representative_field = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, representative_xpath))
        )
        representative_field.send_keys(str(biz_name) if biz_name else "")
        time.sleep(1.5)
        
        # '신청구분'에서 '작업환경측정' 체크박스 클릭
        checkbox_xpath = '//*[@id="contents"]/section/section[2]/div[1]/div/label'
        work_environment_checkbox = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, checkbox_xpath))
        )
        work_environment_checkbox.click()
        time.sleep(1.5)
        
        # '반기 종류 선택' 드롭다운 버튼 클릭 및 선택
        try:
            # 드롭다운 버튼 클릭 (title 또는 id 기준)
            dropdown_button = None
            for btn_xpath in ["//button[@title='반기 종류 선택']", "//*[@id='myBtnmySelect']"]:
                try:
                    dropdown_button = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, btn_xpath))
                    )
                    if dropdown_button:
                        break
                except:
                    continue
            
            if not dropdown_button:
                raise Exception("반기 드롭다운 버튼을 찾을 수 없습니다.")
                
            dropdown_button.click()
            time.sleep(1.5)
            
            # 반기 종류 선택 (상반기 또는 하반기)
            half_option_text = "하반기" if "하반기" in str(half) else "상반기"
            
            # 매칭할 XPath 후보군 정의
            xpath_candidates = [
                f"//button[@title='반기 종류 선택']/following-sibling::ul/li//*[contains(text(), '{half_option_text}')]",
                f"//button[@title='반기 종류 선택']/..//ul/li//*[contains(text(), '{half_option_text}')]",
                f"//*[@id='myBtnmySelect']/following-sibling::ul/li//*[contains(text(), '{half_option_text}')]",
                f"//*[@id='myBtnmySelect']/..//ul/li//*[contains(text(), '{half_option_text}')]",
                f"//ul[@tabindex='0']//*[contains(text(), '{half_option_text}')]",
                f"//*[contains(text(), '{half_option_text}:')]",
                f"//*[contains(text(), '{half_option_text}')]"
            ]
            
            half_option = None
            for xpath in xpath_candidates:
                try:
                    half_option = WebDriverWait(self.driver, 3).until(
                        EC.element_to_be_clickable((By.XPATH, xpath))
                    )
                    if half_option:
                        break
                except:
                    continue
            
            if half_option:
                half_option.click()
                self.update_progress(f"반기 종류 선택 완료: {half_option_text}")
                time.sleep(1.5)
            else:
                raise Exception("반기 옵션 항목을 찾을 수 없습니다.")
                
        except Exception as e:
            self.update_progress(f"반기 종류 선택 중 오류 발생: {str(e)}")
            return "FAIL"
        
        # '희망측정기관' 검색 버튼 클릭
        search_button_xpath = '//*[@id="contents"]/section/section[3]/div/button'
        search_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, search_button_xpath))
        )
        search_button.click()
        time.sleep(1.5)
        
        # 희망측정기관 조회 - 검색 조건 "주소" 입력 필드에 "천안시 서북구 늘푸른3길 22" 입력
        search_input_xpath = '//*[@id="searchText"]'
        search_input = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, search_input_xpath))
        )
        search_input.clear()
        search_input.send_keys("천안시 서북구 늘푸른3길 22")
        time.sleep(2)
        
        # "조회" 버튼 클릭
        search_button_xpath = '/html/body/div[3]/div/section/div/div[1]/button'
        search_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, search_button_xpath))
        )
        search_button.click()
        time.sleep(1.5)  # 조회 버튼 클릭 후 비동기 데이터 갱신을 위해 대기
        
        # '한결' 체크박스 요소 클릭 재시도 루프 (stale element 대비)
        han_gyeol_checkbox_xpath = '//*[@id="MeasurementInstitutionListGrid"]/div/div[1]/div[2]/div[2]/div/div[1]/table/tbody/tr/td[1]/label'
        clicked = False
        for attempt in range(3):
            try:
                # 요소가 클릭 가능할 때까지 대기 (매 시도마다 신규 획득)
                han_gyeol_checkbox = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, han_gyeol_checkbox_xpath))
                )
                # JavaScript를 사용하여 클릭 (더 안정적)
                self.driver.execute_script("arguments[0].click();", han_gyeol_checkbox)
                clicked = True
                break
            except Exception as e:
                self.update_progress(f"'한결' 체크박스 클릭 {attempt + 1}회차 실패 (재시도 중): {str(e)}")
                time.sleep(1.5)
                
        if not clicked:
            raise Exception("'한결' 체크박스를 클릭하지 못했습니다 (stale element 예외 지속).")
        time.sleep(1.5)
        
        # 작업환경 측정기관 조회 "확인" 버튼 클릭
        confirm_button_xpath = '/html/body/div[3]/div/section/footer/button[2]'
        confirm_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, confirm_button_xpath))
        )
        confirm_button.click()
        time.sleep(1.5)
        
        # 담당자 이름 입력
        name_xpath = '//*[@id="name"]'
        name_input = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, name_xpath))
        )
        name_input.send_keys(biz_contact_name)
        time.sleep(1.5)
        
        # 담당자 연락처 입력
        phone_xpath = '//*[@id="contents"]/section/section[6]/div[2]/div/input'
        phone_input = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, phone_xpath))
        )
        phone_input.clear()
        phone_input.send_keys(normalize_contact_phone_suffix(biz_contact_phone))
        time.sleep(1.5)
        
        # 고용보험 근로자수 "검색" 버튼 클릭
        search_button_xpath = '//*[@id="contents"]/section/section[7]/div[1]/div/button'
        search_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, search_button_xpath))
        )
        search_button.click()
        time.sleep(1.5)
        
        # 고용보험 근로자수 확인 (#percentage) - 확인 실패 시 제출 금지
        try:
            percentage_elem = WebDriverWait(self.driver, 8).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '#percentage'))
            )
            percentage_text = percentage_elem.text.strip() or percentage_elem.get_attribute("value")
            employee_result = classify_employee_check(percentage_text)
            if employee_result != "OK":
                self.update_progress(f"  -> 근로자 수 확인 결과 {employee_result}. 자동 신청을 중단합니다.")
                return employee_result
        except Exception as e:
            self.update_progress(f"  -> 근로자 수 확인 실패. 자동 신청을 중단합니다: {str(e)}")
            return "EMPLOYEE_CHECK_FAILED"

        # '작업환경측정 비용지원 환수 동의' 체크박스 클릭
        agreement_checkbox_xpath = '//*[@id="contents"]/section/section[8]/div[2]/label'
        agreement_checkbox = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, agreement_checkbox_xpath))
        )
        agreement_checkbox.click()
        time.sleep(1.5)
        
        # 최종 신청 "완료" 클릭
        final_apply_button_xpath = '//*[@id="contents"]/footer/div[2]/button'
        final_apply_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, final_apply_button_xpath))
        )
        final_apply_button.click()
        time.sleep(1.5)
        
        # "신청 정보 확인 안내" 팝업창 "확인" 클릭
        popup_button_xpath = '//*[@id="app"]/div[7]/section/footer/button[2]'
        popup_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, popup_button_xpath))
        )
        popup_button.click()
        try:
            WebDriverWait(self.driver, 12).until(
                lambda current_driver: has_application_success_marker(
                    current_driver.find_element(By.TAG_NAME, "body").text
                )
            )
        except Exception:
            self.update_progress("  -> 신청 완료 응답을 명확히 확인하지 못했습니다.")
            return "APPLY_RESULT_UNKNOWN"

        self.driver.get(url)
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="contents"]/header/div[2]/button'))
        )
        return "OK"
    
    def check_result(self, year, half):
        """결과 확인 자동화 로직"""
        workbook = None
        try:
            workbook, sheet = self.load_excel_file()
            self.update_progress("엑셀 파일을 성공적으로 불러왔습니다.")
        except Exception as e:
            self.update_progress(f"엑셀 파일을 불러오는 데 실패했습니다: {str(e)}")
            return
        
        # 엑셀 데이터 목록 추출
        business_ids = []
        business_start_nums = []
        ceo_names = []
        row_indices = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            biz_id = row[0].value  # A열: 사업장관리번호
            biz_start_num = row[1].value  # B열: 사업개시번호
            ceo_name = row[3].value  # D열: 대표자명
            h_value = row[7].value # H열: 신청여부
            result_value = row[8].value if len(row) > 8 else None  # I열: 신청결과
            
            # I열에 "대상" 또는 "비대상"이 있으면 건너뛰기
            if result_value in ["대상", "비대상"]:
                continue
            
            # H열에 "50인 이상" 또는 "정보없음"이 있으면 건너뛰기
            if h_value in ["50인 이상", "정보없음"]:
                continue
            
            if biz_id and str(biz_id).strip():
                # 대표자명에서 첫 번째 이름만 추출 (쉼표로 구분된 경우)
                if ceo_name and "," in str(ceo_name):
                    ceo_name = str(ceo_name).split(",")[0].strip()
                elif ceo_name:
                    ceo_name = str(ceo_name).strip()
                else:
                    ceo_name = ""
                
                # 사업개시번호 처리
                if biz_start_num:
                    biz_start_num = str(biz_start_num).strip()
                else:
                    biz_start_num = ""
                
                business_ids.append(str(biz_id).strip())
                business_start_nums.append(biz_start_num)
                ceo_names.append(ceo_name)
                row_indices.append(row_idx)
        
        if not business_ids:
            self.update_progress("처리할 데이터가 없습니다.")
            if workbook:
                workbook.close()
            return
        
        self.update_progress(f"처리할 사업장 수: {len(business_ids)}개")
        
        try:
            # Chrome 드라이버 초기화 (이미 있으면 재사용)
            self.init_webdriver()
            
            url = "https://portal.kosha.or.kr/business-apply-search/health-support/step-stone/cont/sub1"
            self.driver.get(url)
            
            # 웹 페이지 로딩 대기
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="contents"]/article/div[1]/div/figure[1]/div/input'))
            )
        except Exception as e:
            self.update_progress(f"웹 페이지 로딩에 실패했습니다: {str(e)}")
            if workbook:
                workbook.close()
            return
        
        # 각 사업장에 대해 반복 처리
        for idx, (business_id, business_start_num, ceo_name, row_idx) in enumerate(zip(business_ids, business_start_nums, ceo_names, row_indices), 1):
            # 일시 중지 체크
            self.wait_if_paused()
            if not self.is_running:
                break
            
            self.update_progress(f"\n[{idx}/{len(business_ids)}] 사업장관리번호: {business_id} 처리 중...")
            
            try:
                # 입력 필드 초기화
                biz_id_field = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="contents"]/article/div[1]/div/figure[1]/div/input'))
                )
                biz_id_field.clear()
                time.sleep(1.5)
                
                # 사업장관리번호 입력
                biz_id_field.send_keys(business_id)
                time.sleep(0.5)
                
                # 대표자 입력 필드
                ceo_field = self.driver.find_element(By.XPATH, '//*[@id="contents"]/article/div[1]/div/figure[3]/div/input')
                ceo_field.clear()
                time.sleep(1.5)
                
                # 대표자명 입력
                if ceo_name:
                    ceo_field.send_keys(ceo_name)
                time.sleep(1.5)
                
                # 사업개시번호 입력
                try:
                    business_start_field = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, '#contents > article > div.listSearch > div > figure:nth-child(4) > div > input[type=text]'))
                    )
                    business_start_field.clear()
                    time.sleep(1.5)
                    
                    if business_start_num:
                        business_start_field.send_keys(business_start_num)
                    time.sleep(1.5)
                except Exception as e:
                    pass
                
                # 조회 버튼 클릭
                search_button = self.driver.find_element(By.XPATH, '//*[@id="contents"]/article/div[1]/button')
                search_button.click()
                time.sleep(2)
                
                # 결과 대기
                try:
                    WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, '#applicateGrid'))
                    )
                    
                    # 테이블 처리
                    try:
                        table = self.driver.find_element(By.CSS_SELECTOR, '#applicateGrid')
                        rows = table.find_elements(By.TAG_NAME, "tr")[1:]
                        
                        # 지정된 반기(상반기/하반기) 신청결과 찾기
                        result = None
                        for row in rows:
                            cols = row.find_elements(By.TAG_NAME, "td")
                            if len(cols) >= 8:
                                # 테이블 구조에 따라 인덱스 확인 필요
                                # 일반적으로: 연도, 반기, ..., 결과 순서
                                year_text = cols[0].text.strip() if len(cols) > 0 else ""
                                half_text = cols[1].text.strip() if len(cols) > 1 else ""
                                result_text = cols[7].text.strip() if len(cols) > 7 else ""
                                
                                # 신청년도와 반기가 일치하는 경우
                                if year_text == str(year) and half_text == half:
                                    result = result_text
                                    break
                        
                        # 결과를 찾지 못한 경우, 반기만으로도 찾기 시도
                        if not result:
                            for row in rows:
                                cols = row.find_elements(By.TAG_NAME, "td")
                                if len(cols) >= 8:
                                    half_text = cols[1].text.strip() if len(cols) > 1 else ""
                                    result_text = cols[7].text.strip() if len(cols) > 7 else ""
                                    if half_text == half:
                                        result = result_text
                                        break
                        
                        # 결과 업데이트 (I열: 9열)
                        if result:
                            self.update_progress(f"  {half} 신청결과: {result}")
                            sheet.cell(row=row_idx, column=9, value=result)
                        else:
                            self.update_progress(f"  {half} 신청결과를 찾을 수 없음")
                            sheet.cell(row=row_idx, column=9, value="결과 없음")
                            
                    except Exception as e:
                        self.update_progress(f"  테이블 처리 중 오류: {str(e)}")
                        sheet.cell(row=row_idx, column=9, value="테이블 오류")
                        
                except Exception as e:
                    self.update_progress(f"  조회 결과 대기 중 오류: {str(e)}")
                    try:
                        no_data_msg = self.driver.find_element(By.XPATH, '//*[contains(text(), "조회된 내역이 없습니다")]')
                        if no_data_msg:
                            sheet.cell(row=row_idx, column=9, value="결과 없음")
                        else:
                            sheet.cell(row=row_idx, column=9, value="조회 실패")
                    except:
                        sheet.cell(row=row_idx, column=9, value="조회 실패")
                
                # 다음 조회를 위해 초기 페이지로 이동
                self.driver.get(url)
                WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="contents"]/article/div[1]/div/figure[1]/div/input'))
                )
                time.sleep(2)
                
            except Exception as e:
                self.update_progress(f"  처리 중 오류 발생: {str(e)}")
                sheet.cell(row=row_idx, column=9, value="오류 발생")
                try:
                    self.driver.get(url)
                    WebDriverWait(self.driver, 15).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="contents"]/article/div[1]/div/figure[1]/div/input'))
                    )
                    time.sleep(2)
                except:
                    break
        
        try:
            workbook.save(self.excel_file_path)
            self.update_progress("\n결과 확인 작업이 완료되었습니다.")
        except Exception as e:
            self.update_progress(f"엑셀 파일 저장 중 오류: {str(e)}")
        finally:
            if workbook:
                workbook.close()
    
    def run(self):
        """GUI 실행"""
        self.root.mainloop()

if __name__ == '__main__':
    app = HealthProgramAutomation()
    app.run()
